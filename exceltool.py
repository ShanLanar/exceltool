# -*- coding: utf-8 -*-
"""
Daten-Tool (GUI) für:
  1) CSV/XLS/XLSX → Excel (Mehrfachauswahl, jede Datei/Sheet als eigener Reiter)
  2) Blattschutz entfernen (Mehrfachauswahl, .xls wird intern nach .xlsx konvertiert)
  3) Parser für OpenTrans / ORDERS05 / EDIFACT mit einheitlichem Excel-Export (Header + Positionen)

Hinweise:
- Benötigt: pandas, openpyxl, xlrd
- Einheitliches Exportformat:
    * Sheet "Header": Schlüssel/Wert-Paare, inkl. Adressen & Ansprechpartner
    * Sheet "Positionen": Position, Artikelnummer, Beschreibung, Menge, Einheit,
                          Preis, Nettowert, Lieferdatum, Incoterm, Steuer

BUGFIX: Syntaxfehler in analyze_text() – unterminated string literals repariert
        (echte Newlines in Strings ersetzt durch \\n Escape-Sequenzen)

Verbesserungen (siehe Commit-History):
  * Blattschutz-Entfernung ist jetzt namespace-sicher (kein Neuschreiben des
    kompletten XML mehr -> keine beschädigten Dateien)
  * GUI-Aufbau in main() gekapselt -> Modul ist ohne Display importier-/testbar
  * diverse Bugfixes (Encoding-Fallback, leere Zielmappe, Exception-Typ)
"""

from __future__ import annotations

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import csv
import math
import zipfile
import shutil
import tempfile
import threading
import queue
import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Font
import re
from typing import Dict, Any, List, Tuple, Optional

logger = logging.getLogger("exceltool")


# =========================================================
# Thread-Infrastruktur (nicht-blockierendes UI)
# =========================================================

# Wird in main() gesetzt; hier vordefiniert, damit Hintergrund-Threads
# gefahrlos prüfen können, ob das Fenster (noch) existiert.
root = None

_ui_queue: "queue.Queue" = queue.Queue()
_task_running = False
_pump_after_id = None


def _post_ui(fn) -> None:
    """Reiht eine UI-Aktion ein; sie wird im Main-Thread abgearbeitet."""
    _ui_queue.put(fn)


def _pump_ui_queue() -> None:
    """Arbeitet eingereihte UI-Aktionen im Main-Thread ab (periodisch)."""
    global _pump_after_id
    try:
        while True:
            fn = _ui_queue.get_nowait()
            try:
                fn()
            except Exception:
                logger.exception("UI-Update-Fehler")
    except queue.Empty:
        pass
    if root is not None:
        try:
            _pump_after_id = root.after(80, _pump_ui_queue)
        except tk.TclError:
            pass  # Fenster wird gerade geschlossen


def _shutdown() -> None:
    """Sauberes Beenden: laufenden Pump-Timer abbrechen, dann Fenster schließen."""
    global root, _pump_after_id
    if root is None:
        return
    if _pump_after_id is not None:
        try:
            root.after_cancel(_pump_after_id)
        except Exception:
            pass
        _pump_after_id = None
    win, root = root, None
    try:
        win.destroy()
    except Exception:
        pass


def _run_in_background(work) -> bool:
    """
    Führt work() in einem Daemon-Thread aus (nicht-blockierendes UI).
    work() darf Tk-Widgets ausschließlich über _post_ui(...) ansprechen.
    Parallele Tasks werden verhindert (gemeinsame Fortschrittsanzeige/Status).
    Gibt False zurück, wenn bereits ein Task läuft.
    """
    global _task_running
    if _task_running:
        messagebox.showinfo("Bitte warten", "Es läuft bereits eine Verarbeitung.")
        return False
    _task_running = True

    def runner():
        global _task_running
        try:
            work()
        except Exception as e:
            logger.exception("Hintergrund-Task fehlgeschlagen")
            _post_ui(lambda e=e: messagebox.showerror("Fehler", f"Unerwarteter Fehler: {e}"))
        finally:
            _task_running = False

    threading.Thread(target=runner, daemon=True).start()
    return True


# =========================================================
# Hilfsfunktionen
# =========================================================

def safe_str(val) -> str:
    """Konvertiert jeden Wert sicher in String und trimmt."""
    try:
        if val is None:
            return "—"
        s = str(val)
        return s.strip() if s.strip() else "—"
    except Exception:
        return "—"


def escape_excel_formula(s: str) -> str:
    """
    Entschärft CSV-/Formel-Injection: Werte aus Fremdquellen (EDIFACT/XML),
    die mit = + - @ oder einem Steuerzeichen beginnen, könnten von Excel als
    Formel ausgewertet werden. Ein vorangestelltes Apostroph macht sie zu Text
    (Excel blendet das Apostroph aus, der angezeigte Wert bleibt gleich).
    """
    if isinstance(s, str) and s[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


def format_excel(file_path: str) -> None:
    """Formatiert alle Sheets: Kopfzeile fett, Spaltenbreite automatisch (max 60)."""
    try:
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            for col in ws.columns:
                max_len = 0
                try:
                    col_letter = col[0].column_letter
                except Exception:
                    continue
                for cell in col:
                    v = cell.value
                    if v is not None:
                        l = len(str(v))
                        if l > max_len:
                            max_len = l
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        wb.save(file_path)
    except Exception as e:
        logger.warning("Formatierung nicht vollständig möglich: %s", e)


def export_to_excel(header_dict: Dict[str, Any], items_list: List[Dict[str, Any]],
                    doc_id: Optional[str] = None) -> None:
    """Exportiert Header und Positionen in eine Excel-Datei."""
    default_name = f"{doc_id or 'export'}.xlsx"
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        initialfile=default_name,
        title="Export speichern unter",
        filetypes=[("Excel-Datei", "*.xlsx")]
    )
    if not save_path:
        return
    try:
        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            header_rows = [{"Feld": k, "Wert": escape_excel_formula(safe_str(v))}
                           for k, v in header_dict.items()]
            pd.DataFrame(header_rows).to_excel(writer, sheet_name="Header", index=False)
            all_keys = [
                "Position", "Artikelnummer", "Beschreibung", "Menge", "Einheit",
                "Preis", "Nettowert", "Lieferdatum", "Incoterm", "Steuer", "Währung"
            ]
            norm_items = [{k: escape_excel_formula(safe_str(it.get(k))) for k in all_keys}
                          for it in items_list]
            pd.DataFrame(norm_items).to_excel(writer, sheet_name="Positionen", index=False)
        format_excel(save_path)
        messagebox.showinfo("Erfolg", f"Export erfolgreich: {save_path}")
    except Exception as e:
        messagebox.showerror("Fehler", f"Export fehlgeschlagen: {e}")


# =========================================================
# CSV/XLS/XLSX → Excel (Konverter)
# =========================================================

selected_files_convert: List[str] = []


def detect_delimiter(file_path: str) -> str:
    """Erkennt automatisch das Trennzeichen einer CSV."""
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            return dialect.delimiter
    except Exception:
        for delimiter in [',', ';', '\t', '|']:
            try:
                pd.read_csv(file_path, nrows=5, delimiter=delimiter)
                return delimiter
            except Exception:
                continue
        return ','


def read_csv_with_encoding_fallback(path: str, delimiter: str) -> pd.DataFrame:
    for enc in ['utf-8-sig', 'utf-8', 'iso-8859-1', 'cp1252']:
        try:
            return pd.read_csv(path, delimiter=delimiter, encoding=enc, on_bad_lines='skip')
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    return pd.read_csv(path, delimiter=delimiter, on_bad_lines='skip')


def standardize_date_str(value: Any) -> Any:
    try:
        return pd.to_datetime(value, dayfirst=True, errors='raise').strftime('%Y-%m-%d')
    except Exception:
        return value


def sanitize_sheet_name(name: str, used: set) -> str:
    """Excel-konforme Blattnamen (max 31 Zeichen, ohne : \\ / ? * [ ]), unique machen."""
    invalid = r'[:\\/*?\[\]]'
    safe = re.sub(invalid, '_', name)
    safe = safe[:31] if len(safe) > 31 else safe
    base = safe or "Sheet"
    new_name = base
    i = 2
    while new_name in used:
        suffix = f"_{i}"
        new_name = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(new_name)
    return new_name


def select_files_convert(listbox: tk.Listbox, status_label: ttk.Label,
                         btn_convert: ttk.Button) -> None:
    global selected_files_convert
    files = filedialog.askopenfilenames(
        title="Dateien auswählen",
        filetypes=[("CSV/XLS/XLSX", "*.csv *.xls *.xlsx")]
    )
    if not files:
        return
    selected_files_convert = list(files)
    listbox.delete(0, tk.END)
    for f in selected_files_convert:
        listbox.insert(tk.END, f)
    btn_convert.config(state="normal")
    status_label.config(text=f"{len(selected_files_convert)} Datei(en) ausgewählt.")


def convert_files_to_workbook(files: List[str], save_path: str,
                              progress=None, on_file_error=None) -> int:
    """
    Reine Konvertierungslogik (ohne GUI, damit headless testbar):
    Schreibt alle Dateien/Sheets als Reiter in eine Excel-Datei und gibt die
    Anzahl geschriebener Blätter zurück. ``progress(i)`` und
    ``on_file_error(datei, exc)`` sind optionale Callbacks.
    """
    used_names = set()
    written = 0
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        for i, file in enumerate(files, start=1):
            ext = os.path.splitext(file)[1].lower()
            base = os.path.splitext(os.path.basename(file))[0]
            try:
                if ext == ".csv":
                    delimiter = detect_delimiter(file)
                    df = read_csv_with_encoding_fallback(file, delimiter)
                    for col in df.select_dtypes(include=['object']).columns:
                        series = df[col]
                        if series.astype(str).str.contains(r'[/-]', regex=True, na=False).any():
                            df[col] = series.apply(standardize_date_str)
                    numeric_cols = df.select_dtypes(
                        include=['float64', 'int64', 'int32', 'float32']
                    ).columns
                    for col in numeric_cols:
                        df[col] = df[col].apply(
                            lambda x: str(x).replace('.', ',') if pd.notnull(x) else x
                        )
                    sheet_name = sanitize_sheet_name(base, used_names)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    written += 1
                elif ext in [".xls", ".xlsx"]:
                    engine = "openpyxl" if ext == ".xlsx" else "xlrd"
                    sheets = pd.read_excel(file, sheet_name=None, engine=engine)
                    for sh_name, df in sheets.items():
                        combined = f"{base}_{sh_name}"
                        sheet_name = sanitize_sheet_name(combined, used_names)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        written += 1
            except Exception as e:
                if on_file_error:
                    on_file_error(file, e)
            if progress:
                progress(i)

        # Ohne mindestens ein Blatt kann openpyxl die Mappe nicht speichern
        if written == 0:
            pd.DataFrame({"Hinweis": ["Keine konvertierbaren Daten gefunden."]}).to_excel(
                writer, sheet_name="Leer", index=False)

    format_excel(save_path)
    return written


def csv_xls_to_excel(progress_bar: ttk.Progressbar, status_label: ttk.Label) -> None:
    if not selected_files_convert:
        messagebox.showerror("Fehler", "Keine Dateien ausgewählt!")
        return
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        title="Zieldatei speichern unter",
        filetypes=[("Excel-Datei", "*.xlsx")]
    )
    if not save_path:
        return
    files = list(selected_files_convert)
    status_label.config(text="Verarbeite Dateien…")
    progress_bar["maximum"] = len(files)
    progress_bar["value"] = 0

    def work():
        def progress(i):
            _post_ui(lambda i=i: progress_bar.config(value=i))

        def on_file_error(file, e):
            _post_ui(lambda file=file, e=e:
                     messagebox.showerror("Fehler", f"Fehler bei {file}: {e}"))

        written = convert_files_to_workbook(files, save_path, progress, on_file_error)
        if written == 0:
            _post_ui(lambda: status_label.config(text="Keine Daten konvertiert."))
            _post_ui(lambda: messagebox.showwarning(
                "Hinweis", "Es konnten keine Daten konvertiert werden.\n"
                           "Bitte Eingabedateien prüfen."))
        else:
            _post_ui(lambda: status_label.config(text=f"Fertig! Datei gespeichert: {save_path}"))
            _post_ui(lambda: messagebox.showinfo("Erfolg", f"Excel-Datei erstellt: {save_path}"))

    _run_in_background(work)


# =========================================================
# Blattschutz entfernen (Stapellauf)
# =========================================================

selected_files_protect: List[str] = []


def select_files_protect(listbox: tk.Listbox, status_label: ttk.Label,
                         btn_protect: ttk.Button) -> None:
    """Excel-Dateien wählen und Button-Zustand steuern."""
    global selected_files_protect
    files = filedialog.askopenfilenames(
        title="Excel-Dateien auswählen (.xls/.xlsx)",
        filetypes=[("Excel-Dateien", "*.xls *.xlsx")]
    )
    selected_files_protect = list(files) if files else []
    listbox.delete(0, tk.END)
    for f in selected_files_protect:
        listbox.insert(tk.END, f)
    btn_protect.config(state=("normal" if selected_files_protect else "disabled"))
    status_label.config(
        text=f"{len(selected_files_protect)} Excel-Datei(en) für Entschützen ausgewählt."
    )


def _strip_protection_xml(xml_bytes: bytes) -> bytes:
    """
    Entfernt Schutz-Elemente (sheetProtection / workbookProtection / fileSharing)
    rein textuell aus einer OOXML-XML-Datei.

    Bewusst KEIN Re-Parsing/Neuschreiben via ElementTree: dabei würden
    Namespace-Präfixe umbenannt (z. B. mc:Ignorable-Referenzen), was Excel als
    "unlesbaren Inhalt" markiert. Die textuelle Variante lässt die restliche
    Struktur unverändert.
    """
    try:
        text = xml_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return xml_bytes
    for tag in ("sheetProtection", "workbookProtection", "fileSharing"):
        # Leere Elemente (Normalfall), optionaler Namespace-Präfix
        text = re.sub(rf'<(?:\w+:)?{tag}\b[^>]*?/>', '', text)
        # Selten: Element mit explizitem End-Tag
        text = re.sub(rf'<(?:\w+:)?{tag}\b[^>]*?>.*?</(?:\w+:)?{tag}>', '', text,
                      flags=re.DOTALL)
    return text.encode("utf-8")


def entferne_schutz_on_file(xlsx_or_xls_path: str) -> Tuple[bool, str]:
    """
    Entfernt Sheet- und Workbook-Schutz, indem nur die betroffenen XML-Teile der
    OOXML-Datei textuell bereinigt werden; alle übrigen ZIP-Einträge bleiben
    unverändert erhalten. Bei .xls (BIFF) wird zuvor nach .xlsx konvertiert.
    """
    ext = os.path.splitext(xlsx_or_xls_path)[1].lower()
    out_dir = os.path.dirname(xlsx_or_xls_path)
    base = os.path.splitext(os.path.basename(xlsx_or_xls_path))[0]
    neue_datei = os.path.join(out_dir, base + "_entschuetzt.xlsx")

    tmp_dir = tempfile.mkdtemp(prefix="entschuetzt_")
    try:
        if ext == ".xls":
            try:
                source_for_zip = os.path.join(tmp_dir, base + ".xlsx")
                sheets = pd.read_excel(xlsx_or_xls_path, sheet_name=None, engine="xlrd")
                with pd.ExcelWriter(source_for_zip, engine="openpyxl") as writer:
                    for name, df in sheets.items():
                        df.to_excel(writer, sheet_name=str(name)[:31], index=False)
            except Exception as e:
                return False, f"Konvertierung von .xls fehlgeschlagen: {e}"
        elif ext == ".xlsx":
            source_for_zip = xlsx_or_xls_path
        else:
            return False, f"Nicht unterstütztes Format: {ext}"

        # ZIP einlesen, Schutz-Elemente entfernen, neues ZIP schreiben
        with zipfile.ZipFile(source_for_zip, "r") as zin:
            entries = zin.infolist()
            payload = {info.filename: zin.read(info.filename) for info in entries}

        for name in payload:
            if (name.startswith("xl/worksheets/") and name.endswith(".xml")) \
                    or name == "xl/workbook.xml":
                payload[name] = _strip_protection_xml(payload[name])

        with zipfile.ZipFile(neue_datei, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in entries:
                # ZipInfo erhält Zeitstempel/Kompressionstyp des Originals
                zout.writestr(info, payload[info.filename])

        return True, neue_datei
    except Exception as e:
        return False, str(e)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def remove_protection_batch(progress_bar: ttk.Progressbar, status_label: ttk.Label) -> None:
    if not selected_files_protect:
        return
    files = list(selected_files_protect)
    status_label.config(text="Entferne Blattschutz…")
    progress_bar["maximum"] = len(files)
    progress_bar["value"] = 0

    def work():
        results = []
        for i, file in enumerate(files, start=1):
            ok, msg = entferne_schutz_on_file(file)
            results.append((file, ok, msg))
            _post_ui(lambda i=i: progress_bar.config(value=i))
        success = [r for r in results if r[1]]
        fail = [r for r in results if not r[1]]
        info_text = ""
        if success:
            info_text += "Erfolgreich entschützt:\n" + "\n".join(
                [f"- {r[2]}" for r in success]) + "\n\n"
        if fail:
            info_text += "Fehlgeschlagen:\n" + "\n".join(
                [f"- {r[0]}: {r[2]}" for r in fail])
        _post_ui(lambda: status_label.config(text="Entschützen abgeschlossen."))
        _post_ui(lambda info_text=info_text:
                 messagebox.showinfo("Ergebnis", info_text or "Keine Ergebnisse."))

    _run_in_background(work)


# =========================================================
# Parser: OpenTrans / ORDERS05 / EDIFACT
# =========================================================

last_parsed_header: Dict[str, Any] = {}
last_parsed_items: List[Dict[str, Any]] = []
last_doc_id: Optional[str] = None


# -------------------------------
# OpenTrans Parser (XML)
# -------------------------------

def parse_opentrans_xml(raw: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Optional[str]]:
    items: List[Dict[str, Any]] = []
    header: Dict[str, Any] = {}
    try:
        root_xml = ET.fromstring(raw)
    except Exception as e:
        return {"Fehler": f"XML konnte nicht geparst werden: {e}"}, items, None

    def ns_tag(tag: str) -> str:
        if root_xml.tag.startswith("{"):
            ns = root_xml.tag.split("}")[0].strip("{")
            return f"{{{ns}}}{tag}"
        return tag

    def find_text_any(paths: List[str]) -> str:
        for p in paths:
            el = root_xml.find(p)
            if el is not None and el.text:
                txt = el.text.strip()
                if txt:
                    return txt
        return "—"

    doc_id = find_text_any([f".//{ns_tag('ORDER_ID')}", f".//{ns_tag('ORDERNUMBER')}"])
    date = find_text_any([f".//{ns_tag('ORDER_DATE')}", f".//{ns_tag('DATE')}"])
    currency = find_text_any([f".//{ns_tag('CURRENCY')}", f".//{ns_tag('Currency')}"])

    header["Dokumenttyp"] = root_xml.tag.split('}')[-1] if root_xml.tag.startswith("{") else root_xml.tag
    header["Dokument-ID"] = doc_id
    header["Datum"] = date
    header["Währung"] = currency

    buyer_name = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Name')}",
                                f".//{ns_tag('BUYER_PARTY')}/{ns_tag('NAME')}"])
    buyer_id = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('PartyID')}",
                              f".//{ns_tag('BUYER_PARTY')}/{ns_tag('PARTY_ID')}"])
    buyer_street = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('Street')}"])
    buyer_city = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('City')}"])
    buyer_zip = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('Zip')}"])
    buyer_country = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('Country')}"])
    buyer_contact = find_text_any(
        [f".//{ns_tag('BuyerParty')}/{ns_tag('ContactDetails')}/{ns_tag('ContactName')}"])
    buyer_email = find_text_any(
        [f".//{ns_tag('BuyerParty')}/{ns_tag('ContactDetails')}/{ns_tag('Email')}"])

    supplier_name = find_text_any([f".//{ns_tag('SupplierParty')}/{ns_tag('Name')}",
                                   f".//{ns_tag('SUPPLIER_PARTY')}/{ns_tag('NAME')}"])
    supplier_id = find_text_any([f".//{ns_tag('SupplierParty')}/{ns_tag('PartyID')}",
                                 f".//{ns_tag('SUPPLIER_PARTY')}/{ns_tag('PARTY_ID')}"])
    supplier_street = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('Street')}"])
    supplier_city = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('City')}"])
    supplier_zip = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('Zip')}"])
    supplier_country = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('Country')}"])
    supplier_contact = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('ContactDetails')}/{ns_tag('ContactName')}"])
    supplier_email = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('ContactDetails')}/{ns_tag('Email')}"])

    delivery_name = find_text_any([f".//{ns_tag('DeliveryParty')}/{ns_tag('Name')}"])
    delivery_id = find_text_any([f".//{ns_tag('DeliveryParty')}/{ns_tag('PartyID')}"])
    delivery_street = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('Street')}"])
    delivery_city = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('City')}"])
    delivery_zip = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('Zip')}"])
    delivery_country = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('Country')}"])
    delivery_contact = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('ContactDetails')}/{ns_tag('ContactName')}"])
    delivery_email = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('ContactDetails')}/{ns_tag('Email')}"])

    header.update({
        "Käufer.ID": buyer_id, "Käufer.Name": buyer_name, "Käufer.Straße": buyer_street,
        "Käufer.Ort": buyer_city, "Käufer.PLZ": buyer_zip, "Käufer.Land": buyer_country,
        "Käufer.Ansprechpartner": buyer_contact, "Käufer.E-Mail": buyer_email,
        "Lieferant.ID": supplier_id, "Lieferant.Name": supplier_name,
        "Lieferant.Straße": supplier_street,
        "Lieferant.Ort": supplier_city, "Lieferant.PLZ": supplier_zip,
        "Lieferant.Land": supplier_country,
        "Lieferant.Ansprechpartner": supplier_contact, "Lieferant.E-Mail": supplier_email,
        "Lieferadresse.ID": delivery_id, "Lieferadresse.Name": delivery_name,
        "Lieferadresse.Straße": delivery_street,
        "Lieferadresse.Ort": delivery_city, "Lieferadresse.PLZ": delivery_zip,
        "Lieferadresse.Land": delivery_country,
        "Lieferadresse.Ansprechpartner": delivery_contact,
        "Lieferadresse.E-Mail": delivery_email,
    })

    pos_counter = 0
    for item in root_xml.findall(f".//{ns_tag('OrderItem')}"):
        pos_counter += 1

        def find_in_item(paths: List[str]) -> str:
            for p in paths:
                el = item.find(p)
                if el is not None and el.text:
                    t = el.text.strip()
                    if t:
                        return t
            return "—"

        line_no = find_in_item([f".//{ns_tag('LineItemID')}", f".//{ns_tag('LINE_ITEM_ID')}",
                                f".//{ns_tag('LineNumber')}", f".//{ns_tag('LINE_NUMBER')}"])
        pid = find_in_item([f".//{ns_tag('ProductID')}", f".//{ns_tag('PRODUCT_ID')}"])
        name = find_in_item([f".//{ns_tag('ProductName')}", f".//{ns_tag('PRODUCT_NAME')}",
                             f".//{ns_tag('DESCRIPTION_SHORT')}"])
        qty = find_in_item([f".//{ns_tag('Quantity')}", f".//{ns_tag('QUANTITY')}"])
        price = find_in_item([f".//{ns_tag('PriceAmount')}", f".//{ns_tag('PRICE_AMOUNT')}"])
        net = find_in_item([f".//{ns_tag('NetPrice')}"])
        del_date = find_in_item([f".//{ns_tag('DeliveryDate')}", f".//{ns_tag('EDATU')}"])
        incoterm = find_in_item([f".//{ns_tag('Incoterm')}", f".//{ns_tag('LKOND')}"])
        unit = find_in_item([f".//{ns_tag('OrderUnit')}", f".//{ns_tag('ORDER_UNIT')}",
                             f".//{ns_tag('Unit')}", f".//{ns_tag('QuantityUnit')}"])
        tax = find_in_item([f".//{ns_tag('TaxRate')}", f".//{ns_tag('Tax')}",
                            f".//{ns_tag('TAX')}", f".//{ns_tag('VatRate')}"])
        items.append({
            "Position": line_no if line_no != "—" else str(pos_counter),
            "Artikelnummer": pid,
            "Beschreibung": name,
            "Menge": qty,
            "Einheit": unit,
            "Preis": price,
            "Nettowert": net,
            "Lieferdatum": del_date,
            "Incoterm": incoterm,
            "Steuer": tax,
            "Währung": currency,
        })

    return header, items, doc_id if doc_id != "—" else None


# -------------------------------
# ORDERS05 Parser (XML)
# -------------------------------

def parse_orders05_xml(raw: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Optional[str]]:
    items: List[Dict[str, Any]] = []
    header: Dict[str, Any] = {}
    try:
        root_xml = ET.fromstring(raw)
    except Exception as e:
        return {"Fehler": f"XML konnte nicht geparst werden: {e}"}, items, None

    def get_text(node: ET.Element, tag: str, default: str = "—") -> str:
        el = node.find(tag)
        return (el.text.strip() if (el is not None and el.text) else default)

    edi = root_xml.find(".//EDI_DC40")
    docnum = get_text(edi, "DOCNUM", "—") if edi is not None else "—"
    credat = get_text(edi, "CREDAT", "—") if edi is not None else "—"
    cretim = get_text(edi, "CRETIM", "—") if edi is not None else "—"
    mestyp = get_text(edi, "MESTYP", "—") if edi is not None else "—"

    k01 = root_xml.find(".//E1EDK01")
    curcy = get_text(k01, "CURCY", "—") if k01 is not None else "—"
    zterm = get_text(k01, "ZTERM", "—") if k01 is not None else "—"
    bsart = get_text(k01, "BSART", "—") if k01 is not None else "—"
    belnr = get_text(k01, "BELNR", "—") if k01 is not None else "—"

    parties: Dict[str, Dict[str, str]] = {}
    for ka1 in root_xml.findall(".//E1EDKA1"):
        role = get_text(ka1, "PARVW", "")
        if not role:
            continue
        parties[role] = {
            "ID": get_text(ka1, "PARTN", get_text(ka1, "LIFNR", "—")),
            "Name": get_text(ka1, "NAME1", get_text(ka1, "BNAME", "—")),
            "Straße": get_text(ka1, "STRAS", "—"),
            "Ort": get_text(ka1, "ORT01", "—"),
            "PLZ": get_text(ka1, "PSTLZ", "—"),
            "Land": get_text(ka1, "LAND1", "—"),
            "Ansprechpartner": get_text(ka1, "BNAME", "—"),
            "E-Mail": get_text(ka1, "KNREF", "—"),
            "Telefon": get_text(ka1, "TELF1", "—"),
        }

    header.update({
        "Dokumenttyp": "ORDERS05",
        "Dokument-ID": docnum,
        "Datum": credat,
        "Uhrzeit": cretim,
        "Währung": curcy,
        "MESTYP": mestyp,
        "Zahlungsbedingung": zterm,
        "Bestellart": bsart,
        "Belegnummer": belnr,
    })

    def inject_party(role_label: str, role_code: str) -> None:
        p = parties.get(role_code, {})
        header[f"{role_label}.ID"] = p.get("ID", "—")
        header[f"{role_label}.Name"] = p.get("Name", "—")
        header[f"{role_label}.Straße"] = p.get("Straße", "—")
        header[f"{role_label}.Ort"] = p.get("Ort", "—")
        header[f"{role_label}.PLZ"] = p.get("PLZ", "—")
        header[f"{role_label}.Land"] = p.get("Land", "—")
        header[f"{role_label}.Ansprechpartner"] = p.get("Ansprechpartner", "—")
        header[f"{role_label}.E-Mail"] = p.get("E-Mail", "—")
        header[f"{role_label}.Telefon"] = p.get("Telefon", "—")

    inject_party("Käufer", "AG")
    inject_party("Lieferant", "LF")
    inject_party("Lieferadresse", "WE")

    pos_counter = 0
    for p01 in root_xml.findall(".//E1EDP01"):
        pos_counter += 1
        posex = get_text(p01, "POSEX", str(pos_counter))
        menge = get_text(p01, "MENGE", "—")
        uom = get_text(p01, "MENEE", "—")
        vpreis = get_text(p01, "VPREI", "—")
        netwr = get_text(p01, "NETWR", "—")
        prod_id = "—"
        descr = "—"
        for p19 in p01.findall(".//E1EDP19"):
            qualf = get_text(p19, "QUALF", "")
            if qualf == "002":
                prod_id = get_text(p19, "IDTNR", prod_id)
            elif qualf == "001":
                descr = get_text(p19, "KTEXT", descr)
        edatu = "—"
        wmeng = "—"
        for p20 in p01.findall(".//E1EDP20"):
            wmeng = get_text(p20, "WMENG", wmeng)
            edatu = get_text(p20, "EDATU", edatu)
        lkond = "—"
        for p17 in p01.findall(".//E1EDP17"):
            lkond = get_text(p17, "LKOND", lkond)
        items.append({
            "Position": posex,
            "Artikelnummer": prod_id,
            "Beschreibung": descr,
            "Menge": menge,
            "Einheit": uom,
            "Preis": vpreis,
            "Nettowert": netwr,
            "Lieferdatum": edatu,
            "Incoterm": lkond,
            "Steuer": "—",
            "Währung": curcy,
        })

    return header, items, (docnum if docnum != "—" else None)


# -------------------------------
# EDIFACT Parser (Text)
# -------------------------------

def parse_edifact(raw: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], Optional[str]]:
    """Parser für EDIFACT ORDERS D.96A (heuristisch)."""
    component_sep = ':'
    element_sep = '+'
    release_char = '?'
    segment_terminator = "'"

    raw_stripped = raw.strip()
    if raw_stripped.startswith('UNA'):
        try:
            una_line = raw_stripped.splitlines()[0]
            component_sep = una_line[3]
            element_sep = una_line[4]
            release_char = una_line[6]
            segment_terminator = una_line[8]
        except Exception:
            pass

    segments = [s for s in raw.split(segment_terminator) if s.strip()]

    def split_elements(seg: str) -> List[str]:
        elements = []
        current = ''
        i = 0
        while i < len(seg):
            ch = seg[i]
            if ch == release_char:
                i += 1
                if i < len(seg):
                    current += seg[i]
            elif ch == element_sep:
                elements.append(current)
                current = ''
            else:
                current += ch
            i += 1
        elements.append(current)
        return elements

    def get_element(seg: str, idx: int) -> str:
        parts = split_elements(seg)
        return parts[idx] if len(parts) > idx else ''

    def split_components(elem: str) -> List[str]:
        comps = []
        current = ''
        i = 0
        while i < len(elem):
            ch = elem[i]
            if ch == release_char:
                i += 1
                if i < len(elem):
                    current += elem[i]
            elif ch == component_sep:
                comps.append(current)
                current = ''
            else:
                current += ch
            i += 1
        comps.append(current)
        return comps

    header: Dict[str, Any] = {
        "Dokumenttyp": "EDIFACT ORDERS",
        "Dokument-ID": "—",
        "Datum": "—",
        "Währung": "—",
    }

    parties: Dict[str, Dict[str, str]] = {}
    party_contacts: Dict[str, Dict[str, str]] = {}
    items: List[Dict[str, Any]] = []
    current_item: Optional[Dict[str, Any]] = None
    position_counter = 0

    def ensure_party(role: str):
        if role not in parties:
            parties[role] = {
                "ID": "—", "Name": "—", "Straße": "—", "Ort": "—", "PLZ": "—", "Land": "—"
            }
        if role not in party_contacts:
            party_contacts[role] = {"Ansprechpartner": "—", "E-Mail": "—", "Telefon": "—"}

    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        tag = seg[:3].upper()

        if tag == 'BGM':
            header["Dokument-ID"] = get_element(seg, 2) or header["Dokument-ID"]
        elif tag == 'DTM' and current_item is None:
            dtm_elem = get_element(seg, 1)
            comps = split_components(dtm_elem)
            if len(comps) >= 2 and comps[0] == '137':
                header["Datum"] = comps[1]
        elif tag == 'CUX':
            cux_val = get_element(seg, 1)
            comps = split_components(cux_val)
            if len(comps) >= 2:
                header["Währung"] = comps[1]
        elif tag == 'NAD':
            role = get_element(seg, 1)
            ensure_party(role)
            parties[role]["ID"] = get_element(seg, 2) or parties[role]["ID"]
            name_elem = get_element(seg, 4)
            if name_elem:
                name_parts = name_elem.split('::')
                parties[role]["Name"] = name_parts[0] or parties[role]["Name"]
                if len(name_parts) > 1:
                    contact_guess = name_parts[1].split(':')[-1].strip()
                    if contact_guess:
                        party_contacts[role]["Ansprechpartner"] = contact_guess
            parties[role]["Straße"] = get_element(seg, 5) or parties[role]["Straße"]
            parties[role]["Ort"] = get_element(seg, 6) or parties[role]["Ort"]
            parties[role]["PLZ"] = get_element(seg, 8) or parties[role]["PLZ"]
            parties[role]["Land"] = get_element(seg, 9) or parties[role]["Land"]
        elif tag == 'COM':
            val = get_element(seg, 1)
            comps = split_components(val)
            email = comps[0] if comps else ''
            for r in ['BY', 'SU', 'DP']:
                ensure_party(r)
                if party_contacts[r]["E-Mail"] in ("—", ""):
                    party_contacts[r]["E-Mail"] = email or party_contacts[r]["E-Mail"]
        elif tag == 'LIN':
            if current_item:
                items.append(current_item)
            position_counter += 1
            current_item = {
                "Position": str(position_counter),
                "Artikelnummer": "—",
                "Beschreibung": "—",
                "Menge": "—",
                "Einheit": "—",
                "Preis": "—",
                "Nettowert": "—",
                "Lieferdatum": "—",
                "Incoterm": "—",
                "Steuer": "—",
                "Währung": header.get("Währung", "—"),
            }
        elif tag == 'PIA' and current_item:
            val = get_element(seg, 2)
            pid = val.split(':')[0].split('#')[0] if val else ''
            if pid:
                current_item["Artikelnummer"] = pid
        elif tag == 'IMD' and current_item:
            desc = get_element(seg, 3)
            if desc:
                parts = split_components(desc)
                current_item["Beschreibung"] = parts[-1].strip() if parts else desc
        elif tag == 'FTX' and current_item:
            extra = get_element(seg, 4)
            if extra:
                if current_item["Beschreibung"] in ("—", ""):
                    current_item["Beschreibung"] = extra
                else:
                    current_item["Beschreibung"] += " " + extra
        elif tag == 'QTY' and current_item:
            val = get_element(seg, 1)
            comps = split_components(val)
            if len(comps) >= 2:
                current_item["Menge"] = comps[1]
            if len(comps) >= 3:
                current_item["Einheit"] = comps[2]
        elif tag == 'DTM' and current_item:
            val = get_element(seg, 1)
            comps = split_components(val)
            if len(comps) >= 2 and comps[0] in ('2', '10', '35', '64', '94'):
                current_item["Lieferdatum"] = comps[1]
        elif tag == 'PRI' and current_item:
            val = get_element(seg, 1)
            comps = split_components(val)
            if len(comps) >= 2:
                current_item["Preis"] = comps[1]
        elif tag == 'TAX' and current_item:
            try:
                rate_elem = get_element(seg, 5)
                if rate_elem:
                    comps = split_components(rate_elem)
                    if comps:
                        current_item["Steuer"] = comps[-1]
            except Exception:
                pass
        elif tag == 'UNT':
            if current_item:
                items.append(current_item)
                current_item = None

    if current_item:
        items.append(current_item)

    def inject_party(role_label: str, role_code: str) -> None:
        p = parties.get(role_code, {})
        c = party_contacts.get(role_code, {})
        header[f"{role_label}.ID"] = p.get("ID", "—")
        header[f"{role_label}.Name"] = p.get("Name", "—")
        header[f"{role_label}.Straße"] = p.get("Straße", "—")
        header[f"{role_label}.Ort"] = p.get("Ort", "—")
        header[f"{role_label}.PLZ"] = p.get("PLZ", "—")
        header[f"{role_label}.Land"] = p.get("Land", "—")
        header[f"{role_label}.Ansprechpartner"] = c.get("Ansprechpartner", "—")
        header[f"{role_label}.E-Mail"] = c.get("E-Mail", "—")
        header[f"{role_label}.Telefon"] = c.get("Telefon", "—")

    inject_party("Käufer", "BY")
    inject_party("Lieferant", "SU")
    inject_party("Lieferadresse", "DP")

    doc_id = header.get("Dokument-ID")
    return header, items, (doc_id if doc_id != "—" else None)


# -------------------------------
# Analyse-Handler
# -------------------------------

def analyze_text(parser_input: tk.Text, parser_output: tk.Text,
                 btn_export_items: ttk.Button) -> None:
    """
    BUGFIX: Alle parser_output.insert()-Aufrufe nutzen jetzt korrekte \\n-Escape-Sequenzen
    statt echter Newlines im String-Literal (das war der SyntaxError).
    """
    global last_parsed_header, last_parsed_items, last_doc_id
    raw = parser_input.get("1.0", tk.END).strip()
    if not raw:
        messagebox.showerror("Fehler", "Bitte OpenTrans-, ORDERS05- oder EDIFACT-Text einfügen.")
        return
    last_parsed_header = {}
    last_parsed_items = []
    last_doc_id = None

    if raw.startswith("<"):
        try:
            root_xml = ET.fromstring(raw)
            root_name = root_xml.tag.split('}')[-1] if root_xml.tag.startswith("{") else root_xml.tag
        except Exception:
            root_name = ""
        if ("ORDERS05" in raw) or (root_name.upper() in ["ORDERS05", "IDOC"]):
            header, items, doc_id = parse_orders05_xml(raw)
        else:
            header, items, doc_id = parse_opentrans_xml(raw)
    else:
        header, items, doc_id = parse_edifact(raw)

    last_parsed_header = header
    last_parsed_items = items
    last_doc_id = doc_id

    parser_output.config(state="normal")
    parser_output.delete("1.0", tk.END)

    # *** BUGFIX: \n als Escape-Sequenz, nicht als echter Zeilenumbruch im Literal ***
    parser_output.insert(tk.END, "Kopf-Informationen:\n")
    for k, v in header.items():
        parser_output.insert(tk.END, f"  {k}: {safe_str(v)}\n")
    parser_output.insert(tk.END, "\nPositionen:\n")

    if items:
        cols = [
            "Position", "Artikelnummer", "Beschreibung", "Menge", "Einheit",
            "Preis", "Nettowert", "Lieferdatum", "Incoterm", "Steuer", "Währung"
        ]
        for it in items:
            line = " | ".join([safe_str(it.get(c)) for c in cols])
            parser_output.insert(tk.END, f"  - {line}\n")
    else:
        parser_output.insert(tk.END, "  (keine Positionen erkannt)\n")

    parser_output.config(state="disabled")
    btn_export_items.config(state=("normal" if items else "disabled"))


def export_items_to_excel() -> None:
    export_to_excel(last_parsed_header, last_parsed_items, last_doc_id)


def copy_summary_to_clipboard(parser_output: tk.Text) -> None:
    txt = parser_output.get("1.0", tk.END)
    root.clipboard_clear()
    root.clipboard_append(txt)
    messagebox.showinfo("Kopiert", "Zusammenfassung in die Zwischenablage kopiert.")


# =========================================================
# Tab 4: CSV-Splitter (split_csv_2000)
# =========================================================

selected_files_split: List[str] = []


def select_files_split(listbox: tk.Listbox, status_label: ttk.Label,
                       btn_split: ttk.Button) -> None:
    """CSV-Dateien für den Splitter auswählen."""
    global selected_files_split
    files = filedialog.askopenfilenames(
        title="CSV-Dateien auswählen",
        filetypes=[("CSV-Dateien", "*.csv"), ("Alle Dateien", "*.*")]
    )
    selected_files_split = list(files) if files else []
    listbox.delete(0, tk.END)
    for f in selected_files_split:
        listbox.insert(tk.END, f)
    btn_split.config(state=("normal" if selected_files_split else "disabled"))
    status_label.config(text=f"{len(selected_files_split)} CSV-Datei(en) ausgewählt.")


def split_csv_file(file_path: str, chunk_size: int, log_fn) -> int:
    """
    Teilt eine CSV in Blöcke à chunk_size Zeilen (exkl. Header).
    Gibt Anzahl erzeugter Dateien zurück.
    Encoding-Fallback: utf-8 → iso-8859-1 → cp1252.
    """
    header: List[str] = []
    rows: List[List[str]] = []
    for enc in ["utf-8-sig", "utf-8", "iso-8859-1", "cp1252"]:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                header = next(reader, [])
                rows   = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"Konnte {file_path} mit keinem bekannten Encoding lesen.")

    base_name = os.path.splitext(os.path.basename(file_path))[0]
    out_dir   = os.path.dirname(file_path)
    total     = len(rows)
    chunks    = math.ceil(total / chunk_size) if total else 1
    file_count = 0

    for i in range(chunks):
        chunk_rows = rows[i * chunk_size:(i + 1) * chunk_size]
        out_path   = os.path.join(out_dir, f"{base_name}_part{i+1:03d}.csv")
        with open(out_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(chunk_rows)
        log_fn(f"  ✔ Teil {i+1}/{chunks}: {len(chunk_rows)} Zeilen → {os.path.basename(out_path)}")
        file_count += 1

    return file_count


def run_split(chunk_size_var: tk.IntVar, progress_bar: ttk.Progressbar,
              status_label: ttk.Label, log_widget: tk.Text) -> None:
    """Startet den Splitter für alle ausgewählten Dateien."""
    if not selected_files_split:
        return

    chunk_size = chunk_size_var.get()
    if chunk_size < 1:
        messagebox.showerror("Fehler", "Zeilenzahl pro Teil muss ≥ 1 sein.")
        return

    files = list(selected_files_split)

    def log(msg):
        def _do():
            log_widget.config(state="normal")
            log_widget.insert("end", msg + "\n")
            log_widget.see("end")
            log_widget.config(state="disabled")
        _post_ui(_do)

    progress_bar.config(value=0, maximum=len(files))
    status_label.config(text="Splitte CSV-Dateien…")

    def work():
        total_parts = 0
        for i, file_path in enumerate(files, 1):
            log(f"\n→ {os.path.basename(file_path)}")
            try:
                parts = split_csv_file(file_path, chunk_size, log)
                total_parts += parts
            except Exception as e:
                log(f"  ✖ Fehler: {e}")
            _post_ui(lambda i=i: progress_bar.config(value=i))
        _post_ui(lambda total_parts=total_parts:
                 status_label.config(text=f"Fertig: {total_parts} Teildatei(en) erzeugt."))
        _post_ui(lambda total_parts=total_parts: messagebox.showinfo(
            "CSV-Splitter",
            f"{total_parts} Teildatei(en) erzeugt.\nAusgabe im jeweiligen Quellordner."))

    _run_in_background(work)


# =========================================================
# Tab 5: XML → CSV (etc_partner_wandeln)
# =========================================================

def parse_parameter_xml(xml_text: str) -> Tuple[List[str], List[str]]:
    """
    Liest ein XML-Fragment mit <PARAMETER DISPLAYNAME="...">Wert</PARAMETER>
    und gibt (header, values) zurück.
    Toleriert fehlende Wurzel – wrapping wird automatisch versucht.
    """
    # Ggf. künstliche Wurzel ergänzen
    text = xml_text.strip()
    if not text.startswith("<"):
        raise ValueError("Kein gültiges XML erkannt.")
    try:
        root_el = ET.fromstring(text)
    except ET.ParseError:
        try:
            root_el = ET.fromstring(f"<ROOT>{text}</ROOT>")
        except ET.ParseError as e:
            raise ValueError(f"XML konnte nicht geparst werden: {e}")

    params = root_el.findall(".//PARAMETER")
    if not params:
        # Fallback: direkte Kinder
        params = list(root_el)
    if not params:
        raise ValueError("Keine PARAMETER-Elemente im XML gefunden.")

    header = [p.get("DISPLAYNAME", p.tag) for p in params]
    values = [p.text or "" for p in params]
    return header, values


def xml_to_csv_convert(xml_input: tk.Text, log_widget: tk.Text,
                       sep_var: tk.StringVar) -> None:
    """Liest XML aus dem Eingabefeld und speichert als CSV."""
    xml_text = xml_input.get("1.0", tk.END).strip()
    if not xml_text:
        messagebox.showwarning("Hinweis", "Bitte XML in das Eingabefeld einfügen.")
        return

    def log(msg):
        log_widget.config(state="normal")
        log_widget.insert("end", msg + "\n")
        log_widget.see("end")
        log_widget.config(state="disabled")

    try:
        header, values = parse_parameter_xml(xml_text)
    except ValueError as e:
        messagebox.showerror("Fehler", str(e))
        return

    sep = sep_var.get() or ";"
    save_path = filedialog.asksaveasfilename(
        title="CSV speichern unter",
        defaultextension=".csv",
        initialfile="partner_export.csv",
        filetypes=[("CSV-Datei", "*.csv")]
    )
    if not save_path:
        return

    try:
        with open(save_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=sep)
            writer.writerow(header)
            writer.writerow(values)
        log(f"✔ {len(header)} Felder exportiert → {save_path}")
        log(f"  Felder: {', '.join(header[:8])}{'…' if len(header) > 8 else ''}")
        messagebox.showinfo("Erfolg", f"CSV gespeichert:\n{save_path}\n\n{len(header)} Felder.")
    except Exception as e:
        messagebox.showerror("Fehler", f"Speichern fehlgeschlagen: {e}")


def paste_from_clipboard(xml_input: tk.Text) -> None:
    """Holt Inhalt aus Zwischenablage und setzt ihn ins Eingabefeld."""
    try:
        content = root.clipboard_get()
        xml_input.delete("1.0", tk.END)
        xml_input.insert("1.0", content)
    except tk.TclError:
        messagebox.showwarning("Zwischenablage", "Zwischenablage ist leer oder enthält keinen Text.")


def preview_xml(xml_input: tk.Text, preview_widget: tk.Text) -> None:
    """Zeigt eine Tabellenvorschau der geparsten Felder."""
    xml_text = xml_input.get("1.0", tk.END).strip()
    if not xml_text:
        return
    try:
        header, values = parse_parameter_xml(xml_text)
    except ValueError as e:
        preview_widget.config(state="normal")
        preview_widget.delete("1.0", tk.END)
        preview_widget.insert("1.0", f"Fehler: {e}")
        preview_widget.config(state="disabled")
        return

    preview_widget.config(state="normal")
    preview_widget.delete("1.0", tk.END)
    preview_widget.insert("end", f"{len(header)} Felder erkannt:\n\n")
    col_w = 28
    for h, v in zip(header, values):
        line = f"  {h[:col_w]:<{col_w}}  {v[:60]}\n"
        preview_widget.insert("end", line)
    preview_widget.config(state="disabled")


# =========================================================
# GUI Aufbau
# =========================================================

def _scrolled(parent, factory, pack_kw, horizontal=False):
    """
    Bettet ein Listbox-/Text-Widget mit Scrollbar(s) in einen eigenen Rahmen ein.
    ``factory(frame)`` erzeugt das Widget; ``pack_kw`` platziert den Rahmen im
    übergeordneten Container (wie zuvor das Widget selbst).
    """
    frame = ttk.Frame(parent)
    frame.pack(**pack_kw)
    widget = factory(frame)
    vsb = ttk.Scrollbar(frame, orient="vertical", command=widget.yview)
    widget.configure(yscrollcommand=vsb.set)
    widget.grid(row=0, column=0, sticky="nsew")
    vsb.grid(row=0, column=1, sticky="ns")
    if horizontal:
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=widget.xview)
        widget.configure(xscrollcommand=hsb.set)
        hsb.grid(row=1, column=0, sticky="ew")
    frame.rowconfigure(0, weight=1)
    frame.columnconfigure(0, weight=1)
    return widget


def main():
    global root
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    root = tk.Tk()
    root.title("Excel-Toolbox: CSV→Excel | Blattschutz | Parser | CSV-Splitter | XML→CSV")
    root.geometry("1000x740")
    root.minsize(820, 600)

    tab_control = ttk.Notebook(root)
    tab_csv     = ttk.Frame(tab_control)
    tab_protect = ttk.Frame(tab_control)
    tab_parser  = ttk.Frame(tab_control)
    tab_split   = ttk.Frame(tab_control)
    tab_xml2csv = ttk.Frame(tab_control)

    tab_control.add(tab_csv,     text="📂 CSV/XLS/XLSX → Excel")
    tab_control.add(tab_protect, text="🔓 Blattschutz entfernen")
    tab_control.add(tab_parser,  text="📄 OpenTrans/ORDERS05/EDIFACT")
    tab_control.add(tab_split,   text="✂ CSV-Splitter")
    tab_control.add(tab_xml2csv, text="🔄 XML → CSV")
    tab_control.pack(expand=1, fill="both")

    # --- Tab 1: Konverter ---
    frame_csv_top = ttk.Frame(tab_csv)
    frame_csv_top.pack(fill="x", padx=10, pady=10)
    ttk.Label(frame_csv_top,
              text="Wähle CSV/XLS/XLSX-Dateien und füge alle als Reiter in eine Excel-Datei zusammen."
              ).pack(side="left")

    convert_listbox = _scrolled(
        tab_csv, lambda p: tk.Listbox(p, height=12),
        dict(fill="both", expand=True, padx=10, pady=5))

    frame_csv_bottom = ttk.Frame(tab_csv)
    frame_csv_bottom.pack(fill="x", padx=10, pady=10)

    status_label_global = ttk.Label(root, text="", foreground="blue")
    progress_bar_global = ttk.Progressbar(root, length=400, mode="determinate")

    btn_run_convert = ttk.Button(
        frame_csv_bottom, text="Zusammenführen & Speichern…",
        command=lambda: csv_xls_to_excel(progress_bar_global, status_label_global),
        state="disabled"
    )
    btn_select_convert = ttk.Button(
        frame_csv_bottom, text="Dateien auswählen…",
        command=lambda: select_files_convert(convert_listbox, status_label_global, btn_run_convert)
    )
    btn_select_convert.pack(side="left")
    btn_run_convert.pack(side="left", padx=10)

    # --- Tab 2: Schutz entfernen ---
    frame_protect_top = ttk.Frame(tab_protect)
    frame_protect_top.pack(fill="x", padx=10, pady=10)
    ttk.Label(frame_protect_top,
              text="Wähle Excel-Dateien (.xls/.xlsx), um Blattschutz/Arbeitsmappenschutz zu entfernen."
              ).pack(side="left")

    protect_listbox = _scrolled(
        tab_protect, lambda p: tk.Listbox(p, height=12),
        dict(fill="both", expand=True, padx=10, pady=5))

    frame_protect_bottom = ttk.Frame(tab_protect)
    frame_protect_bottom.pack(fill="x", padx=10, pady=10)

    btn_run_protect = ttk.Button(
        frame_protect_bottom, text="Blattschutz entfernen (Stapellauf)",
        command=lambda: remove_protection_batch(progress_bar_global, status_label_global),
        state="disabled"
    )
    btn_select_protect = ttk.Button(
        frame_protect_bottom, text="Excel-Dateien auswählen…",
        command=lambda: select_files_protect(protect_listbox, status_label_global, btn_run_protect)
    )
    btn_select_protect.pack(side="left")
    btn_run_protect.pack(side="left", padx=10)

    # --- Tab 3: Parser ---
    frame_parser_top = ttk.Frame(tab_parser)
    frame_parser_top.pack(fill="x", padx=10, pady=10)
    ttk.Label(frame_parser_top,
              text="OpenTrans-XML, SAP IDoc ORDERS05-XML oder EDIFACT-Rohtext einfügen und analysieren."
              ).pack(side="left")

    parser_input = _scrolled(
        tab_parser, lambda p: tk.Text(p, height=16),
        dict(fill="both", expand=True, padx=10, pady=5))

    frame_parser_actions = ttk.Frame(tab_parser)
    frame_parser_actions.pack(fill="x", padx=10, pady=5)

    btn_export_items = ttk.Button(
        frame_parser_actions, text="Export (Header + Positionen) → Excel…",
        command=export_items_to_excel, state="disabled"
    )
    btn_analyze = ttk.Button(
        frame_parser_actions, text="Analysieren",
        command=lambda: analyze_text(parser_input, parser_output, btn_export_items)
    )
    btn_copy = ttk.Button(
        frame_parser_actions, text="Zusammenfassung kopieren",
        command=lambda: copy_summary_to_clipboard(parser_output)
    )
    btn_analyze.pack(side="left")
    btn_copy.pack(side="left", padx=10)
    btn_export_items.pack(side="left", padx=10)

    parser_output = _scrolled(
        tab_parser, lambda p: tk.Text(p, height=18, state="disabled"),
        dict(fill="both", expand=True, padx=10, pady=5))

    # --- Tab 4: CSV-Splitter ---
    ttk.Label(tab_split,
              text="Teilt große CSV-Dateien in gleichmäßige Blöcke auf.\n"
                   "Jeder Block enthält die Kopfzeile. Ausgabe im Quellordner der Originaldatei."
              ).pack(anchor="w", padx=10, pady=(10, 4))

    frm_split_opts = ttk.LabelFrame(tab_split, text="Einstellungen")
    frm_split_opts.pack(fill="x", padx=10, pady=4)
    ttk.Label(frm_split_opts, text="Zeilen pro Datei:").grid(row=0, column=0, sticky="w", padx=8, pady=6)
    split_chunk_var = tk.IntVar(value=2000)
    ttk.Spinbox(frm_split_opts, from_=100, to=100000, increment=500,
                textvariable=split_chunk_var, width=10).grid(row=0, column=1, sticky="w", padx=6, pady=6)
    ttk.Label(frm_split_opts, text="(Standard: 2000 – passend für viele Shop-Importe)").grid(
        row=0, column=2, sticky="w", padx=6)

    split_listbox = _scrolled(
        tab_split, lambda p: tk.Listbox(p, height=7),
        dict(fill="both", expand=False, padx=10, pady=4))

    frm_split_log = ttk.LabelFrame(tab_split, text="Protokoll")
    frm_split_log.pack(fill="both", expand=True, padx=10, pady=4)
    split_log = _scrolled(
        frm_split_log, lambda p: tk.Text(p, height=8, state="disabled"),
        dict(fill="both", expand=True, padx=5, pady=5))

    frm_split_btn = ttk.Frame(tab_split)
    frm_split_btn.pack(fill="x", padx=10, pady=6)

    btn_run_split = ttk.Button(
        frm_split_btn, text="Splitten",
        command=lambda: run_split(split_chunk_var, progress_bar_global, status_label_global, split_log),
        state="disabled"
    )
    btn_select_split = ttk.Button(
        frm_split_btn, text="CSV-Dateien auswählen…",
        command=lambda: select_files_split(split_listbox, status_label_global, btn_run_split)
    )
    btn_select_split.pack(side="left")
    btn_run_split.pack(side="left", padx=10)

    # --- Tab 5: XML → CSV ---
    ttk.Label(tab_xml2csv,
              text="Liest ein XML-Fragment mit <PARAMETER DISPLAYNAME=\"...\">Wert</PARAMETER>\n"
                   "und exportiert es als CSV. Typisch für ETC-Partner-Daten aus der Zwischenablage."
              ).pack(anchor="w", padx=10, pady=(10, 4))

    frm_xml_opts = ttk.LabelFrame(tab_xml2csv, text="Einstellungen")
    frm_xml_opts.pack(fill="x", padx=10, pady=4)
    ttk.Label(frm_xml_opts, text="Trennzeichen:").grid(row=0, column=0, sticky="w", padx=8, pady=6)
    xml_sep_var = tk.StringVar(value=";")
    sep_box = ttk.Combobox(frm_xml_opts, values=[";", ",", "\t", "|"],
                           textvariable=xml_sep_var, width=6, state="normal")
    sep_box.grid(row=0, column=1, sticky="w", padx=6, pady=6)
    ttk.Label(frm_xml_opts, text="(Standard: Semikolon)").grid(row=0, column=2, sticky="w", padx=6)

    frm_xml_input = ttk.LabelFrame(tab_xml2csv, text="XML-Eingabe (einfügen oder Zwischenablage)")
    frm_xml_input.pack(fill="both", expand=True, padx=10, pady=4)
    xml_input = _scrolled(
        frm_xml_input, lambda p: tk.Text(p, height=10, wrap="none"),
        dict(fill="both", expand=True, padx=5, pady=5), horizontal=True)

    frm_xml_preview = ttk.LabelFrame(tab_xml2csv, text="Vorschau erkannter Felder")
    frm_xml_preview.pack(fill="both", expand=True, padx=10, pady=4)
    xml_preview = _scrolled(
        frm_xml_preview, lambda p: tk.Text(p, height=7, state="disabled",
                                           font=("Courier New", 9)),
        dict(fill="both", expand=True, padx=5, pady=5), horizontal=True)

    frm_xml_btn = ttk.Frame(tab_xml2csv)
    frm_xml_btn.pack(fill="x", padx=10, pady=6)
    ttk.Button(frm_xml_btn, text="📋 Aus Zwischenablage",
               command=lambda: paste_from_clipboard(xml_input)).pack(side="left")
    ttk.Button(frm_xml_btn, text="🔍 Vorschau",
               command=lambda: preview_xml(xml_input, xml_preview)).pack(side="left", padx=8)
    ttk.Button(frm_xml_btn, text="💾 Als CSV speichern…",
               command=lambda: xml_to_csv_convert(xml_input, xml_preview, xml_sep_var)).pack(side="left")
    ttk.Button(frm_xml_btn, text="🗑 Eingabe leeren",
               command=lambda: xml_input.delete("1.0", tk.END)).pack(side="right")

    # --- Globaler Status ---
    status_frame = ttk.Frame(root)
    status_frame.pack(fill="x", padx=10, pady=5)
    status_label_global.pack(side="left")
    progress_bar_global.pack(side="right")

    # Sauberes Beenden (Pump-Timer abbrechen) beim Schließen des Fensters
    root.protocol("WM_DELETE_WINDOW", _shutdown)
    # Hintergrund-Tasks melden UI-Updates über diese Pumpe zurück
    _pump_ui_queue()
    root.mainloop()


# =========================================================
# Start
# =========================================================
if __name__ == "__main__":
    main()
