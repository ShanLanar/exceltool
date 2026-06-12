# -*- coding: utf-8 -*-
"""
core.py - GUI-freie Logik von exceltool (headless test- und wiederverwendbar).

Enthaelt Parser (OpenTrans / ORDERS05 / EDIFACT / PARAMETER-XML), den
CSV/XLS/XLSX-zu-Excel-Konverter, den CSV-Splitter sowie die namespace-sichere
Blattschutz-Entfernung. Keine tkinter-Abhaengigkeit.
"""

from __future__ import annotations

import csv
import logging
import math
import os
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

logger = logging.getLogger("exceltool.core")


def safe_str(val) -> str:
    """Konvertiert jeden Wert sicher in String und trimmt."""
    try:
        if val is None:
            return "—"
        s = str(val)
        return s.strip() if s.strip() else "—"
    except Exception:
        return "—"


def escape_excel_formula(s: str) -> str:
    """
    Entschärft CSV-/Formel-Injection: Werte aus Fremdquellen (EDIFACT/XML),
    die mit = + - @ oder einem Steuerzeichen beginnen, könnten von Excel als
    Formel ausgewertet werden. Ein vorangestelltes Apostroph macht sie zu Text
    (Excel blendet das Apostroph aus, der angezeigte Wert bleibt gleich).
    """
    if isinstance(s, str) and s[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


def format_excel(file_path: str) -> None:
    """Formatiert alle Sheets: Kopfzeile fett, Spaltenbreite automatisch (max 60)."""
    try:
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            for col in ws.columns:
                max_len = 0
                try:
                    col_letter = col[0].column_letter
                except Exception:
                    continue
                for cell in col:
                    v = cell.value
                    if v is not None:
                        length = len(str(v))
                        if length > max_len:
                            max_len = length
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        wb.save(file_path)
    except Exception as e:
        logger.warning("Formatierung nicht vollständig möglich: %s", e)


def detect_delimiter(file_path: str) -> str:
    """Erkennt automatisch das Trennzeichen einer CSV."""
    try:
        with open(file_path, encoding='utf-8-sig') as f:
            sample = f.read(4096)
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
            return dialect.delimiter
    except Exception:
        for delimiter in [',', ';', '\t', '|']:
            try:
                pd.read_csv(file_path, nrows=5, delimiter=delimiter)
                return delimiter
            except Exception:
                continue
        return ','


def read_csv_with_encoding_fallback(path: str, delimiter: str) -> pd.DataFrame:
    for enc in ['utf-8-sig', 'utf-8', 'iso-8859-1', 'cp1252']:
        try:
            return pd.read_csv(path, delimiter=delimiter, encoding=enc, on_bad_lines='skip')
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    return pd.read_csv(path, delimiter=delimiter, on_bad_lines='skip')


def standardize_date_str(value: Any) -> Any:
    try:
        return pd.to_datetime(value, dayfirst=True, errors='raise').strftime('%Y-%m-%d')
    except Exception:
        return value


def sanitize_sheet_name(name: str, used: set) -> str:
    """Excel-konforme Blattnamen (max 31 Zeichen, ohne : \\ / ? * [ ]), unique machen."""
    invalid = r'[:\\/*?\[\]]'
    safe = re.sub(invalid, '_', name)
    safe = safe[:31] if len(safe) > 31 else safe
    base = safe or "Sheet"
    new_name = base
    i = 2
    while new_name in used:
        suffix = f"_{i}"
        new_name = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(new_name)
    return new_name


def convert_files_to_workbook(files: list[str], save_path: str,
                              progress=None, on_file_error=None) -> int:
    """
    Reine Konvertierungslogik (ohne GUI, damit headless testbar):
    Schreibt alle Dateien/Sheets als Reiter in eine Excel-Datei und gibt die
    Anzahl geschriebener Blätter zurück. ``progress(i)`` und
    ``on_file_error(datei, exc)`` sind optionale Callbacks.
    """
    used_names = set()
    written = 0
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        for i, file in enumerate(files, start=1):
            ext = os.path.splitext(file)[1].lower()
            base = os.path.splitext(os.path.basename(file))[0]
            try:
                if ext == ".csv":
                    delimiter = detect_delimiter(file)
                    df = read_csv_with_encoding_fallback(file, delimiter)
                    for col in df.select_dtypes(include=['object']).columns:
                        series = df[col]
                        if series.astype(str).str.contains(r'[/-]', regex=True, na=False).any():
                            df[col] = series.apply(standardize_date_str)
                    numeric_cols = df.select_dtypes(
                        include=['float64', 'int64', 'int32', 'float32']
                    ).columns
                    for col in numeric_cols:
                        df[col] = df[col].apply(
                            lambda x: str(x).replace('.', ',') if pd.notnull(x) else x
                        )
                    sheet_name = sanitize_sheet_name(base, used_names)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    written += 1
                elif ext in [".xls", ".xlsx"]:
                    engine = "openpyxl" if ext == ".xlsx" else "xlrd"
                    sheets = pd.read_excel(file, sheet_name=None, engine=engine)
                    for sh_name, df in sheets.items():
                        combined = f"{base}_{sh_name}"
                        sheet_name = sanitize_sheet_name(combined, used_names)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        written += 1
            except Exception as e:
                if on_file_error:
                    on_file_error(file, e)
            if progress:
                progress(i)

        # Ohne mindestens ein Blatt kann openpyxl die Mappe nicht speichern
        if written == 0:
            pd.DataFrame({"Hinweis": ["Keine konvertierbaren Daten gefunden."]}).to_excel(
                writer, sheet_name="Leer", index=False)

    format_excel(save_path)
    return written


def _strip_protection_xml(xml_bytes: bytes) -> bytes:
    """
    Entfernt Schutz-Elemente (sheetProtection / workbookProtection / fileSharing)
    rein textuell aus einer OOXML-XML-Datei.

    Bewusst KEIN Re-Parsing/Neuschreiben via ElementTree: dabei würden
    Namespace-Präfixe umbenannt (z. B. mc:Ignorable-Referenzen), was Excel als
    "unlesbaren Inhalt" markiert. Die textuelle Variante lässt die restliche
    Struktur unverändert.
    """
    try:
        text = xml_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return xml_bytes
    for tag in ("sheetProtection", "workbookProtection", "fileSharing"):
        # Leere Elemente (Normalfall), optionaler Namespace-Präfix
        text = re.sub(rf'<(?:\w+:)?{tag}\b[^>]*?/>', '', text)
        # Selten: Element mit explizitem End-Tag
        text = re.sub(rf'<(?:\w+:)?{tag}\b[^>]*?>.*?</(?:\w+:)?{tag}>', '', text,
                      flags=re.DOTALL)
    return text.encode("utf-8")


def entferne_schutz_on_file(xlsx_or_xls_path: str) -> tuple[bool, str]:
    """
    Entfernt Sheet- und Workbook-Schutz, indem nur die betroffenen XML-Teile der
    OOXML-Datei textuell bereinigt werden; alle übrigen ZIP-Einträge bleiben
    unverändert erhalten. Bei .xls (BIFF) wird zuvor nach .xlsx konvertiert.
    """
    ext = os.path.splitext(xlsx_or_xls_path)[1].lower()
    out_dir = os.path.dirname(xlsx_or_xls_path)
    base = os.path.splitext(os.path.basename(xlsx_or_xls_path))[0]
    neue_datei = os.path.join(out_dir, base + "_entschuetzt.xlsx")

    tmp_dir = tempfile.mkdtemp(prefix="entschuetzt_")
    try:
        if ext == ".xls":
            try:
                source_for_zip = os.path.join(tmp_dir, base + ".xlsx")
                sheets = pd.read_excel(xlsx_or_xls_path, sheet_name=None, engine="xlrd")
                with pd.ExcelWriter(source_for_zip, engine="openpyxl") as writer:
                    for name, df in sheets.items():
                        df.to_excel(writer, sheet_name=str(name)[:31], index=False)
            except Exception as e:
                return False, f"Konvertierung von .xls fehlgeschlagen: {e}"
        elif ext == ".xlsx":
            source_for_zip = xlsx_or_xls_path
        else:
            return False, f"Nicht unterstütztes Format: {ext}"

        # ZIP einlesen, Schutz-Elemente entfernen, neues ZIP schreiben
        with zipfile.ZipFile(source_for_zip, "r") as zin:
            entries = zin.infolist()
            payload = {info.filename: zin.read(info.filename) for info in entries}

        for name in payload:
            if (name.startswith("xl/worksheets/") and name.endswith(".xml")) \
                    or name == "xl/workbook.xml":
                payload[name] = _strip_protection_xml(payload[name])

        with zipfile.ZipFile(neue_datei, "w", zipfile.ZIP_DEFLATED) as zout:
            for info in entries:
                # ZipInfo erhält Zeitstempel/Kompressionstyp des Originals
                zout.writestr(info, payload[info.filename])

        return True, neue_datei
    except Exception as e:
        return False, str(e)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def parse_opentrans_xml(raw: str) -> tuple[dict[str, Any], list[dict[str, Any]], str | None]:
    items: list[dict[str, Any]] = []
    header: dict[str, Any] = {}
    try:
        root_xml = ET.fromstring(raw)
    except Exception as e:
        return {"Fehler": f"XML konnte nicht geparst werden: {e}"}, items, None

    def ns_tag(tag: str) -> str:
        if root_xml.tag.startswith("{"):
            ns = root_xml.tag.split("}")[0].strip("{")
            return f"{{{ns}}}{tag}"
        return tag

    def find_text_any(paths: list[str]) -> str:
        for p in paths:
            el = root_xml.find(p)
            if el is not None and el.text:
                txt = el.text.strip()
                if txt:
                    return txt
        return "—"

    doc_id = find_text_any([f".//{ns_tag('ORDER_ID')}", f".//{ns_tag('ORDERNUMBER')}"])
    date = find_text_any([f".//{ns_tag('ORDER_DATE')}", f".//{ns_tag('DATE')}"])
    currency = find_text_any([f".//{ns_tag('CURRENCY')}", f".//{ns_tag('Currency')}"])

    header["Dokumenttyp"] = root_xml.tag.split('}')[-1] if root_xml.tag.startswith("{") else root_xml.tag
    header["Dokument-ID"] = doc_id
    header["Datum"] = date
    header["Währung"] = currency

    buyer_name = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Name')}",
                                f".//{ns_tag('BUYER_PARTY')}/{ns_tag('NAME')}"])
    buyer_id = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('PartyID')}",
                              f".//{ns_tag('BUYER_PARTY')}/{ns_tag('PARTY_ID')}"])
    buyer_street = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('Street')}"])
    buyer_city = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('City')}"])
    buyer_zip = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('Zip')}"])
    buyer_country = find_text_any([f".//{ns_tag('BuyerParty')}/{ns_tag('Address')}/{ns_tag('Country')}"])
    buyer_contact = find_text_any(
        [f".//{ns_tag('BuyerParty')}/{ns_tag('ContactDetails')}/{ns_tag('ContactName')}"])
    buyer_email = find_text_any(
        [f".//{ns_tag('BuyerParty')}/{ns_tag('ContactDetails')}/{ns_tag('Email')}"])

    supplier_name = find_text_any([f".//{ns_tag('SupplierParty')}/{ns_tag('Name')}",
                                   f".//{ns_tag('SUPPLIER_PARTY')}/{ns_tag('NAME')}"])
    supplier_id = find_text_any([f".//{ns_tag('SupplierParty')}/{ns_tag('PartyID')}",
                                 f".//{ns_tag('SUPPLIER_PARTY')}/{ns_tag('PARTY_ID')}"])
    supplier_street = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('Street')}"])
    supplier_city = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('City')}"])
    supplier_zip = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('Zip')}"])
    supplier_country = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('Address')}/{ns_tag('Country')}"])
    supplier_contact = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('ContactDetails')}/{ns_tag('ContactName')}"])
    supplier_email = find_text_any(
        [f".//{ns_tag('SupplierParty')}/{ns_tag('ContactDetails')}/{ns_tag('Email')}"])

    delivery_name = find_text_any([f".//{ns_tag('DeliveryParty')}/{ns_tag('Name')}"])
    delivery_id = find_text_any([f".//{ns_tag('DeliveryParty')}/{ns_tag('PartyID')}"])
    delivery_street = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('Street')}"])
    delivery_city = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('City')}"])
    delivery_zip = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('Zip')}"])
    delivery_country = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('Address')}/{ns_tag('Country')}"])
    delivery_contact = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('ContactDetails')}/{ns_tag('ContactName')}"])
    delivery_email = find_text_any(
        [f".//{ns_tag('DeliveryParty')}/{ns_tag('ContactDetails')}/{ns_tag('Email')}"])

    header.update({
        "Käufer.ID": buyer_id, "Käufer.Name": buyer_name, "Käufer.Straße": buyer_street,
        "Käufer.Ort": buyer_city, "Käufer.PLZ": buyer_zip, "Käufer.Land": buyer_country,
        "Käufer.Ansprechpartner": buyer_contact, "Käufer.E-Mail": buyer_email,
        "Lieferant.ID": supplier_id, "Lieferant.Name": supplier_name,
        "Lieferant.Straße": supplier_street,
        "Lieferant.Ort": supplier_city, "Lieferant.PLZ": supplier_zip,
        "Lieferant.Land": supplier_country,
        "Lieferant.Ansprechpartner": supplier_contact, "Lieferant.E-Mail": supplier_email,
        "Lieferadresse.ID": delivery_id, "Lieferadresse.Name": delivery_name,
        "Lieferadresse.Straße": delivery_street,
        "Lieferadresse.Ort": delivery_city, "Lieferadresse.PLZ": delivery_zip,
        "Lieferadresse.Land": delivery_country,
        "Lieferadresse.Ansprechpartner": delivery_contact,
        "Lieferadresse.E-Mail": delivery_email,
    })

    pos_counter = 0
    for item in root_xml.findall(f".//{ns_tag('OrderItem')}"):
        pos_counter += 1

        def find_in_item(paths: list[str], node=item) -> str:
            for p in paths:
                el = node.find(p)
                if el is not None and el.text:
                    t = el.text.strip()
                    if t:
                        return t
            return "—"

        line_no = find_in_item([f".//{ns_tag('LineItemID')}", f".//{ns_tag('LINE_ITEM_ID')}",
                                f".//{ns_tag('LineNumber')}", f".//{ns_tag('LINE_NUMBER')}"])
        pid = find_in_item([f".//{ns_tag('ProductID')}", f".//{ns_tag('PRODUCT_ID')}"])
        name = find_in_item([f".//{ns_tag('ProductName')}", f".//{ns_tag('PRODUCT_NAME')}",
                             f".//{ns_tag('DESCRIPTION_SHORT')}"])
        qty = find_in_item([f".//{ns_tag('Quantity')}", f".//{ns_tag('QUANTITY')}"])
        price = find_in_item([f".//{ns_tag('PriceAmount')}", f".//{ns_tag('PRICE_AMOUNT')}"])
        net = find_in_item([f".//{ns_tag('NetPrice')}"])
        del_date = find_in_item([f".//{ns_tag('DeliveryDate')}", f".//{ns_tag('EDATU')}"])
        incoterm = find_in_item([f".//{ns_tag('Incoterm')}", f".//{ns_tag('LKOND')}"])
        unit = find_in_item([f".//{ns_tag('OrderUnit')}", f".//{ns_tag('ORDER_UNIT')}",
                             f".//{ns_tag('Unit')}", f".//{ns_tag('QuantityUnit')}"])
        tax = find_in_item([f".//{ns_tag('TaxRate')}", f".//{ns_tag('Tax')}",
                            f".//{ns_tag('TAX')}", f".//{ns_tag('VatRate')}"])
        items.append({
            "Position": line_no if line_no != "—" else str(pos_counter),
            "Artikelnummer": pid,
            "Beschreibung": name,
            "Menge": qty,
            "Einheit": unit,
            "Preis": price,
            "Nettowert": net,
            "Lieferdatum": del_date,
            "Incoterm": incoterm,
            "Steuer": tax,
            "Währung": currency,
        })

    return header, items, doc_id if doc_id != "—" else None


def parse_orders05_xml(raw: str) -> tuple[dict[str, Any], list[dict[str, Any]], str | None]:
    items: list[dict[str, Any]] = []
    header: dict[str, Any] = {}
    try:
        root_xml = ET.fromstring(raw)
    except Exception as e:
        return {"Fehler": f"XML konnte nicht geparst werden: {e}"}, items, None

    def get_text(node: ET.Element, tag: str, default: str = "—") -> str:
        el = node.find(tag)
        return (el.text.strip() if (el is not None and el.text) else default)

    edi = root_xml.find(".//EDI_DC40")
    docnum = get_text(edi, "DOCNUM", "—") if edi is not None else "—"
    credat = get_text(edi, "CREDAT", "—") if edi is not None else "—"
    cretim = get_text(edi, "CRETIM", "—") if edi is not None else "—"
    mestyp = get_text(edi, "MESTYP", "—") if edi is not None else "—"

    k01 = root_xml.find(".//E1EDK01")
    curcy = get_text(k01, "CURCY", "—") if k01 is not None else "—"
    zterm = get_text(k01, "ZTERM", "—") if k01 is not None else "—"
    bsart = get_text(k01, "BSART", "—") if k01 is not None else "—"
    belnr = get_text(k01, "BELNR", "—") if k01 is not None else "—"

    parties: dict[str, dict[str, str]] = {}
    for ka1 in root_xml.findall(".//E1EDKA1"):
        role = get_text(ka1, "PARVW", "")
        if not role:
            continue
        parties[role] = {
            "ID": get_text(ka1, "PARTN", get_text(ka1, "LIFNR", "—")),
            "Name": get_text(ka1, "NAME1", get_text(ka1, "BNAME", "—")),
            "Straße": get_text(ka1, "STRAS", "—"),
            "Ort": get_text(ka1, "ORT01", "—"),
            "PLZ": get_text(ka1, "PSTLZ", "—"),
            "Land": get_text(ka1, "LAND1", "—"),
            "Ansprechpartner": get_text(ka1, "BNAME", "—"),
            "E-Mail": get_text(ka1, "KNREF", "—"),
            "Telefon": get_text(ka1, "TELF1", "—"),
        }

    header.update({
        "Dokumenttyp": "ORDERS05",
        "Dokument-ID": docnum,
        "Datum": credat,
        "Uhrzeit": cretim,
        "Währung": curcy,
        "MESTYP": mestyp,
        "Zahlungsbedingung": zterm,
        "Bestellart": bsart,
        "Belegnummer": belnr,
    })

    def inject_party(role_label: str, role_code: str) -> None:
        p = parties.get(role_code, {})
        header[f"{role_label}.ID"] = p.get("ID", "—")
        header[f"{role_label}.Name"] = p.get("Name", "—")
        header[f"{role_label}.Straße"] = p.get("Straße", "—")
        header[f"{role_label}.Ort"] = p.get("Ort", "—")
        header[f"{role_label}.PLZ"] = p.get("PLZ", "—")
        header[f"{role_label}.Land"] = p.get("Land", "—")
        header[f"{role_label}.Ansprechpartner"] = p.get("Ansprechpartner", "—")
        header[f"{role_label}.E-Mail"] = p.get("E-Mail", "—")
        header[f"{role_label}.Telefon"] = p.get("Telefon", "—")

    inject_party("Käufer", "AG")
    inject_party("Lieferant", "LF")
    inject_party("Lieferadresse", "WE")

    pos_counter = 0
    for p01 in root_xml.findall(".//E1EDP01"):
        pos_counter += 1
        posex = get_text(p01, "POSEX", str(pos_counter))
        menge = get_text(p01, "MENGE", "—")
        uom = get_text(p01, "MENEE", "—")
        vpreis = get_text(p01, "VPREI", "—")
        netwr = get_text(p01, "NETWR", "—")
        prod_id = "—"
        descr = "—"
        for p19 in p01.findall(".//E1EDP19"):
            qualf = get_text(p19, "QUALF", "")
            if qualf == "002":
                prod_id = get_text(p19, "IDTNR", prod_id)
            elif qualf == "001":
                descr = get_text(p19, "KTEXT", descr)
        edatu = "—"
        wmeng = "—"
        for p20 in p01.findall(".//E1EDP20"):
            wmeng = get_text(p20, "WMENG", wmeng)
            edatu = get_text(p20, "EDATU", edatu)
        lkond = "—"
        for p17 in p01.findall(".//E1EDP17"):
            lkond = get_text(p17, "LKOND", lkond)
        items.append({
            "Position": posex,
            "Artikelnummer": prod_id,
            "Beschreibung": descr,
            "Menge": menge,
            "Einheit": uom,
            "Preis": vpreis,
            "Nettowert": netwr,
            "Lieferdatum": edatu,
            "Incoterm": lkond,
            "Steuer": "—",
            "Währung": curcy,
        })

    return header, items, (docnum if docnum != "—" else None)


def parse_edifact(raw: str) -> tuple[dict[str, Any], list[dict[str, Any]], str | None]:
    """Parser für EDIFACT ORDERS D.96A (heuristisch)."""
    component_sep = ':'
    element_sep = '+'
    release_char = '?'
    segment_terminator = "'"

    raw_stripped = raw.strip()
    if raw_stripped.startswith('UNA'):
        try:
            una_line = raw_stripped.splitlines()[0]
            component_sep = una_line[3]
            element_sep = una_line[4]
            release_char = una_line[6]
            segment_terminator = una_line[8]
        except Exception:
            pass

    segments = [s for s in raw.split(segment_terminator) if s.strip()]

    def split_elements(seg: str) -> list[str]:
        elements = []
        current = ''
        i = 0
        while i < len(seg):
            ch = seg[i]
            if ch == release_char:
                i += 1
                if i < len(seg):
                    current += seg[i]
            elif ch == element_sep:
                elements.append(current)
                current = ''
            else:
                current += ch
            i += 1
        elements.append(current)
        return elements

    def get_element(seg: str, idx: int) -> str:
        parts = split_elements(seg)
        return parts[idx] if len(parts) > idx else ''

    def split_components(elem: str) -> list[str]:
        comps = []
        current = ''
        i = 0
        while i < len(elem):
            ch = elem[i]
            if ch == release_char:
                i += 1
                if i < len(elem):
                    current += elem[i]
            elif ch == component_sep:
                comps.append(current)
                current = ''
            else:
                current += ch
            i += 1
        comps.append(current)
        return comps

    header: dict[str, Any] = {
        "Dokumenttyp": "EDIFACT ORDERS",
        "Dokument-ID": "—",
        "Datum": "—",
        "Währung": "—",
    }

    parties: dict[str, dict[str, str]] = {}
    party_contacts: dict[str, dict[str, str]] = {}
    items: list[dict[str, Any]] = []
    current_item: dict[str, Any] | None = None
    position_counter = 0

    def ensure_party(role: str):
        if role not in parties:
            parties[role] = {
                "ID": "—", "Name": "—", "Straße": "—", "Ort": "—", "PLZ": "—", "Land": "—"
            }
        if role not in party_contacts:
            party_contacts[role] = {"Ansprechpartner": "—", "E-Mail": "—", "Telefon": "—"}

    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        tag = seg[:3].upper()

        if tag == 'BGM':
            header["Dokument-ID"] = get_element(seg, 2) or header["Dokument-ID"]
        elif tag == 'DTM' and current_item is None:
            dtm_elem = get_element(seg, 1)
            comps = split_components(dtm_elem)
            if len(comps) >= 2 and comps[0] == '137':
                header["Datum"] = comps[1]
        elif tag == 'CUX':
            cux_val = get_element(seg, 1)
            comps = split_components(cux_val)
            if len(comps) >= 2:
                header["Währung"] = comps[1]
        elif tag == 'NAD':
            role = get_element(seg, 1)
            ensure_party(role)
            parties[role]["ID"] = get_element(seg, 2) or parties[role]["ID"]
            name_elem = get_element(seg, 4)
            if name_elem:
                name_parts = name_elem.split('::')
                parties[role]["Name"] = name_parts[0] or parties[role]["Name"]
                if len(name_parts) > 1:
                    contact_guess = name_parts[1].split(':')[-1].strip()
                    if contact_guess:
                        party_contacts[role]["Ansprechpartner"] = contact_guess
            parties[role]["Straße"] = get_element(seg, 5) or parties[role]["Straße"]
            parties[role]["Ort"] = get_element(seg, 6) or parties[role]["Ort"]
            parties[role]["PLZ"] = get_element(seg, 8) or parties[role]["PLZ"]
            parties[role]["Land"] = get_element(seg, 9) or parties[role]["Land"]
        elif tag == 'COM':
            val = get_element(seg, 1)
            comps = split_components(val)
            email = comps[0] if comps else ''
            for r in ['BY', 'SU', 'DP']:
                ensure_party(r)
                if party_contacts[r]["E-Mail"] in ("—", ""):
                    party_contacts[r]["E-Mail"] = email or party_contacts[r]["E-Mail"]
        elif tag == 'LIN':
            if current_item:
                items.append(current_item)
            position_counter += 1
            current_item = {
                "Position": str(position_counter),
                "Artikelnummer": "—",
                "Beschreibung": "—",
                "Menge": "—",
                "Einheit": "—",
                "Preis": "—",
                "Nettowert": "—",
                "Lieferdatum": "—",
                "Incoterm": "—",
                "Steuer": "—",
                "Währung": header.get("Währung", "—"),
            }
        elif tag == 'PIA' and current_item:
            val = get_element(seg, 2)
            pid = val.split(':')[0].split('#')[0] if val else ''
            if pid:
                current_item["Artikelnummer"] = pid
        elif tag == 'IMD' and current_item:
            desc = get_element(seg, 3)
            if desc:
                parts = split_components(desc)
                current_item["Beschreibung"] = parts[-1].strip() if parts else desc
        elif tag == 'FTX' and current_item:
            extra = get_element(seg, 4)
            if extra:
                if current_item["Beschreibung"] in ("—", ""):
                    current_item["Beschreibung"] = extra
                else:
                    current_item["Beschreibung"] += " " + extra
        elif tag == 'QTY' and current_item:
            val = get_element(seg, 1)
            comps = split_components(val)
            if len(comps) >= 2:
                current_item["Menge"] = comps[1]
            if len(comps) >= 3:
                current_item["Einheit"] = comps[2]
        elif tag == 'DTM' and current_item:
            val = get_element(seg, 1)
            comps = split_components(val)
            if len(comps) >= 2 and comps[0] in ('2', '10', '35', '64', '94'):
                current_item["Lieferdatum"] = comps[1]
        elif tag == 'PRI' and current_item:
            val = get_element(seg, 1)
            comps = split_components(val)
            if len(comps) >= 2:
                current_item["Preis"] = comps[1]
        elif tag == 'TAX' and current_item:
            try:
                rate_elem = get_element(seg, 5)
                if rate_elem:
                    comps = split_components(rate_elem)
                    if comps:
                        current_item["Steuer"] = comps[-1]
            except Exception:
                pass
        elif tag == 'UNT':
            if current_item:
                items.append(current_item)
                current_item = None

    if current_item:
        items.append(current_item)

    def inject_party(role_label: str, role_code: str) -> None:
        p = parties.get(role_code, {})
        c = party_contacts.get(role_code, {})
        header[f"{role_label}.ID"] = p.get("ID", "—")
        header[f"{role_label}.Name"] = p.get("Name", "—")
        header[f"{role_label}.Straße"] = p.get("Straße", "—")
        header[f"{role_label}.Ort"] = p.get("Ort", "—")
        header[f"{role_label}.PLZ"] = p.get("PLZ", "—")
        header[f"{role_label}.Land"] = p.get("Land", "—")
        header[f"{role_label}.Ansprechpartner"] = c.get("Ansprechpartner", "—")
        header[f"{role_label}.E-Mail"] = c.get("E-Mail", "—")
        header[f"{role_label}.Telefon"] = c.get("Telefon", "—")

    inject_party("Käufer", "BY")
    inject_party("Lieferant", "SU")
    inject_party("Lieferadresse", "DP")

    doc_id = header.get("Dokument-ID")
    return header, items, (doc_id if doc_id != "—" else None)


def split_csv_file(file_path: str, chunk_size: int, log_fn) -> int:
    """
    Teilt eine CSV in Blöcke à chunk_size Zeilen (exkl. Header).
    Gibt Anzahl erzeugter Dateien zurück.
    Encoding-Fallback: utf-8 → iso-8859-1 → cp1252.
    """
    header: list[str] = []
    rows: list[list[str]] = []
    for enc in ["utf-8-sig", "utf-8", "iso-8859-1", "cp1252"]:
        try:
            with open(file_path, encoding=enc, newline="") as f:
                reader = csv.reader(f)
                header = next(reader, [])
                rows   = list(reader)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"Konnte {file_path} mit keinem bekannten Encoding lesen.")

    base_name = os.path.splitext(os.path.basename(file_path))[0]
    out_dir   = os.path.dirname(file_path)
    total     = len(rows)
    chunks    = math.ceil(total / chunk_size) if total else 1
    file_count = 0

    for i in range(chunks):
        chunk_rows = rows[i * chunk_size:(i + 1) * chunk_size]
        out_path   = os.path.join(out_dir, f"{base_name}_part{i+1:03d}.csv")
        with open(out_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(chunk_rows)
        log_fn(f"  ✔ Teil {i+1}/{chunks}: {len(chunk_rows)} Zeilen → {os.path.basename(out_path)}")
        file_count += 1

    return file_count


def parse_parameter_xml(xml_text: str) -> tuple[list[str], list[str]]:
    """
    Liest ein XML-Fragment mit <PARAMETER DISPLAYNAME="...">Wert</PARAMETER>
    und gibt (header, values) zurück.
    Toleriert fehlende Wurzel – wrapping wird automatisch versucht.
    """
    # Ggf. künstliche Wurzel ergänzen
    text = xml_text.strip()
    if not text.startswith("<"):
        raise ValueError("Kein gültiges XML erkannt.")
    try:
        root_el = ET.fromstring(text)
    except ET.ParseError:
        try:
            root_el = ET.fromstring(f"<ROOT>{text}</ROOT>")
        except ET.ParseError as e:
            raise ValueError(f"XML konnte nicht geparst werden: {e}") from e

    params = root_el.findall(".//PARAMETER")
    if not params:
        # Fallback: direkte Kinder
        params = list(root_el)
    if not params:
        raise ValueError("Keine PARAMETER-Elemente im XML gefunden.")

    header = [p.get("DISPLAYNAME", p.tag) for p in params]
    values = [p.text or "" for p in params]
    return header, values
