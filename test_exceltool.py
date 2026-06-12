# -*- coding: utf-8 -*-
"""
Regressionstests für exceltool.py (reine Logik, ohne GUI).

Lauf:  python3 test_exceltool.py
       (oder mit pytest:  pytest test_exceltool.py)

tkinter wird nur gestubbt, wenn es nicht installiert ist (Headless-/CI-Umgebung).
Der GUI-Aufbau liegt in main() und wird beim bloßen Import NICHT ausgeführt.
"""

import os
import sys
import time
import types
import tempfile
import zipfile

# --- tkinter ggf. stubben, damit das Modul auch headless importierbar ist ---
import importlib.util

if importlib.util.find_spec("tkinter") is None:  # pragma: no cover - umgebungsabhängig
    _stub = types.ModuleType("tkinter")
    for _sub in ("filedialog", "messagebox", "ttk"):
        _m = types.ModuleType(f"tkinter.{_sub}")
        setattr(_stub, _sub, _m)
        sys.modules[f"tkinter.{_sub}"] = _m
    sys.modules["tkinter"] = _stub

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import exceltool as et  # noqa: E402


# ----------------------------------------------------------------------------
# kleine Test-Infrastruktur (ohne externe Abhängigkeit)
# ----------------------------------------------------------------------------
_failures = []


def check(cond, msg):
    if cond:
        print(f"  ok   - {msg}")
    else:
        print(f"  FAIL - {msg}")
        _failures.append(msg)


# ----------------------------------------------------------------------------
def test_sanitize_sheet_name():
    print("test_sanitize_sheet_name")
    used = set()
    a = et.sanitize_sheet_name("Mappe/1", used)
    check(a == "Mappe_1", f"ungültige Zeichen ersetzt -> {a!r}")
    b = et.sanitize_sheet_name("Mappe/1", used)
    check(b != a and b.startswith("Mappe_1"), f"Duplikat eindeutig gemacht -> {b!r}")
    long = et.sanitize_sheet_name("X" * 50, used)
    check(len(long) <= 31, f"max. 31 Zeichen -> len={len(long)}")


def test_parse_edifact():
    print("test_parse_edifact")
    segs = [
        "UNB+UNOC:3+S+R+240101:1200+1",
        "UNH+1+ORDERS:D:96A:UN",
        "BGM+220+4711+9",
        "DTM+137:20240101:102",
        "CUX+2:EUR:9",
        "NAD+BY+++ACME GmbH+Hauptstr 1+Berlin++10115+DE",
        "NAD+SU+++Lieferant AG",
        "LIN+1++ART-1:BP",
        "PIA+1+ART-1:SA",
        "IMD+F++:::Widget Pro",
        "QTY+21:5:PCE",
        "PRI+AAA:9.99",
        "DTM+2:20240115:102",
        "UNT+11+1",
        "UNZ+1+1",
    ]
    raw = "UNA:+.? '" + "".join(s + "'" for s in segs)
    header, items, doc_id = et.parse_edifact(raw)
    check(doc_id == "4711", f"Dokument-ID -> {doc_id!r}")
    check(header.get("Währung") == "EUR", f"Währung -> {header.get('Währung')!r}")
    check(header.get("Datum") == "20240101", f"Datum -> {header.get('Datum')!r}")
    check(header.get("Käufer.Name") == "ACME GmbH", f"Käufer.Name -> {header.get('Käufer.Name')!r}")
    check(header.get("Käufer.PLZ") == "10115", f"Käufer.PLZ -> {header.get('Käufer.PLZ')!r}")
    check(len(items) == 1, f"genau eine Position -> {len(items)}")
    it = items[0]
    check(it["Artikelnummer"] == "ART-1", f"Artikelnummer -> {it['Artikelnummer']!r}")
    check(it["Beschreibung"] == "Widget Pro", f"Beschreibung -> {it['Beschreibung']!r}")
    check(it["Menge"] == "5", f"Menge -> {it['Menge']!r}")
    check(it["Einheit"] == "PCE", f"Einheit -> {it['Einheit']!r}")
    check(it["Preis"] == "9.99", f"Preis -> {it['Preis']!r}")
    check(it["Lieferdatum"] == "20240115", f"Lieferdatum -> {it['Lieferdatum']!r}")
    check(it["Währung"] == "EUR", f"Währung (Position) -> {it['Währung']!r}")


def test_parse_orders05_xml():
    print("test_parse_orders05_xml")
    raw = """<ORDERS05><IDOC>
        <EDI_DC40><DOCNUM>0815</DOCNUM><CREDAT>20240101</CREDAT>
            <CRETIM>120000</CRETIM><MESTYP>ORDERS</MESTYP></EDI_DC40>
        <E1EDK01><CURCY>EUR</CURCY><BELNR>BN1</BELNR></E1EDK01>
        <E1EDKA1><PARVW>AG</PARVW><PARTN>C1</PARTN><NAME1>ACME</NAME1>
            <ORT01>Berlin</ORT01><PSTLZ>10115</PSTLZ></E1EDKA1>
        <E1EDP01><POSEX>10</POSEX><MENGE>5</MENGE><MENEE>ST</MENEE>
            <VPREI>9.99</VPREI><NETWR>49.95</NETWR>
            <E1EDP19><QUALF>002</QUALF><IDTNR>ART-9</IDTNR></E1EDP19>
            <E1EDP19><QUALF>001</QUALF><KTEXT>Widget</KTEXT></E1EDP19>
            <E1EDP20><WMENG>5</WMENG><EDATU>20240115</EDATU></E1EDP20>
        </E1EDP01>
    </IDOC></ORDERS05>"""
    header, items, doc_id = et.parse_orders05_xml(raw)
    check(doc_id == "0815", f"Dokument-ID -> {doc_id!r}")
    check(header.get("Währung") == "EUR", f"Währung -> {header.get('Währung')!r}")
    check(header.get("Käufer.Name") == "ACME", f"Käufer.Name -> {header.get('Käufer.Name')!r}")
    check(len(items) == 1, f"genau eine Position -> {len(items)}")
    it = items[0]
    check(it["Position"] == "10", f"Position -> {it['Position']!r}")
    check(it["Artikelnummer"] == "ART-9", f"Artikelnummer -> {it['Artikelnummer']!r}")
    check(it["Beschreibung"] == "Widget", f"Beschreibung -> {it['Beschreibung']!r}")
    check(it["Menge"] == "5", f"Menge -> {it['Menge']!r}")
    check(it["Einheit"] == "ST", f"Einheit -> {it['Einheit']!r}")
    check(it["Lieferdatum"] == "20240115", f"Lieferdatum -> {it['Lieferdatum']!r}")


def test_parse_opentrans_xml():
    print("test_parse_opentrans_xml")
    raw = """<ORDER>
        <ORDER_ID>4711</ORDER_ID>
        <ORDER_DATE>2024-01-01</ORDER_DATE>
        <CURRENCY>EUR</CURRENCY>
        <ORDER_ITEM_LIST>
            <OrderItem>
                <ProductID>ART-1</ProductID>
                <ProductName>Widget</ProductName>
                <Quantity>5</Quantity>
                <PriceAmount>9.99</PriceAmount>
            </OrderItem>
        </ORDER_ITEM_LIST>
    </ORDER>"""
    header, items, doc_id = et.parse_opentrans_xml(raw)
    check(doc_id == "4711", f"Dokument-ID -> {doc_id!r}")
    check(header.get("Währung") == "EUR", f"Währung -> {header.get('Währung')!r}")
    check(len(items) == 1, f"genau eine Position -> {len(items)}")
    it = items[0]
    check(it["Artikelnummer"] == "ART-1", f"Artikelnummer -> {it['Artikelnummer']!r}")
    check(it["Beschreibung"] == "Widget", f"Beschreibung -> {it['Beschreibung']!r}")
    check(it["Menge"] == "5", f"Menge -> {it['Menge']!r}")


def test_parse_parameter_xml():
    print("test_parse_parameter_xml")
    wrapped = '<ROOT><PARAMETER DISPLAYNAME="Name">ACME</PARAMETER>' \
              '<PARAMETER DISPLAYNAME="Ort">Berlin</PARAMETER></ROOT>'
    h, v = et.parse_parameter_xml(wrapped)
    check(h == ["Name", "Ort"], f"Header -> {h}")
    check(v == ["ACME", "Berlin"], f"Werte -> {v}")
    # Fragment ohne Wurzel -> automatisches Wrapping
    frag = '<PARAMETER DISPLAYNAME="A">1</PARAMETER><PARAMETER DISPLAYNAME="B">2</PARAMETER>'
    h2, v2 = et.parse_parameter_xml(frag)
    check(h2 == ["A", "B"] and v2 == ["1", "2"], f"Fragment-Wrapping -> {h2}, {v2}")


def test_split_csv_file():
    print("test_split_csv_file")
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "data.csv")
        with open(path, "w", encoding="utf-8", newline="") as f:
            f.write("a,b\n")
            for i in range(5):
                f.write(f"{i},{i*2}\n")
        logs = []
        count = et.split_csv_file(path, chunk_size=2, log_fn=logs.append)
        check(count == 3, f"5 Zeilen / 2 -> 3 Teile (={count})")
        parts = sorted(p for p in os.listdir(d) if "_part" in p)
        check(len(parts) == 3, f"3 Teildateien erzeugt -> {parts}")
        with open(os.path.join(d, parts[0]), encoding="utf-8") as f:
            first = f.readline().strip()
        check(first == "a,b", f"Kopfzeile in Teildatei -> {first!r}")


def test_strip_protection_xml():
    print("test_strip_protection_xml")
    xml = (
        b'<?xml version="1.0"?>'
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        b'mc:Ignorable="x14ac">'
        b'<sheetData/>'
        b'<sheetProtection algorithmName="SHA-512" hashValue="abc" sheet="1" objects="1"/>'
        b'</worksheet>'
    )
    out = et._strip_protection_xml(xml).decode("utf-8")
    check("sheetProtection" not in out, "sheetProtection entfernt")
    check("mc:Ignorable" in out, "mc:Ignorable-Attribut erhalten (kein Namespace-Mangling)")
    check("<sheetData/>" in out, "übrige Struktur erhalten")


def test_entferne_schutz_integration():
    print("test_entferne_schutz_integration")
    from openpyxl import Workbook, load_workbook
    from openpyxl.workbook.protection import WorkbookProtection

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "geschuetzt.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws.protection.sheet = True
        ws.protection.password = "secret"
        wb.security = WorkbookProtection(workbookPassword="secret", lockStructure=True)
        wb.save(path)

        # Vorbedingung: Schutz ist wirklich enthalten
        with zipfile.ZipFile(path) as z:
            ws_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
            wb_xml = z.read("xl/workbook.xml").decode("utf-8")
        check("sheetProtection" in ws_xml, "Vorbedingung: sheetProtection vorhanden")
        check("workbookProtection" in wb_xml, "Vorbedingung: workbookProtection vorhanden")

        ok, out = et.entferne_schutz_on_file(path)
        check(ok, f"Entschützen erfolgreich -> {out}")
        check(os.path.exists(out), "Ausgabedatei existiert")

        with zipfile.ZipFile(out) as z:
            ws_xml2 = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
            wb_xml2 = z.read("xl/workbook.xml").decode("utf-8")
        check("sheetProtection" not in ws_xml2, "sheetProtection entfernt")
        check("workbookProtection" not in wb_xml2, "workbookProtection entfernt")

        wb2 = load_workbook(out)
        check(wb2.active["A1"].value == "hello", "Zellinhalt erhalten")
        check(not wb2.active.protection.sheet, "Blattschutz deaktiviert")


def test_convert_files_to_workbook():
    print("test_convert_files_to_workbook")
    from openpyxl import Workbook, load_workbook
    with tempfile.TemporaryDirectory() as d:
        csv1 = os.path.join(d, "komma.csv")
        with open(csv1, "w", encoding="utf-8", newline="") as f:
            f.write("a,b\n1,2\n3,4\n")
        csv2 = os.path.join(d, "semikolon.csv")
        with open(csv2, "w", encoding="utf-8", newline="") as f:
            f.write("x;y\n5;6\n")
        xlsx_in = os.path.join(d, "mappe.xlsx")
        wb = Workbook()
        wb.active["A1"] = "z"
        wb.save(xlsx_in)

        out = os.path.join(d, "out.xlsx")
        errors = []
        progress = []
        written = et.convert_files_to_workbook(
            [csv1, csv2, xlsx_in], out,
            progress=progress.append,
            on_file_error=lambda f, e: errors.append((f, e)),
        )
        check(written == 3, f"3 Blätter geschrieben -> {written}")
        check(errors == [], f"keine Datei-Fehler -> {errors}")
        check(progress == [1, 2, 3], f"Fortschritt 1..3 gemeldet -> {progress}")
        names = load_workbook(out).sheetnames
        check(len(names) == 3, f"3 Reiter in Zieldatei -> {names}")


def test_convert_empty_creates_placeholder():
    print("test_convert_empty_creates_placeholder")
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as d:
        bad = os.path.join(d, "egal.txt")
        with open(bad, "w", encoding="utf-8") as f:
            f.write("kein tabellenformat")
        out = os.path.join(d, "leer.xlsx")
        written = et.convert_files_to_workbook([bad], out)
        check(written == 0, f"nichts konvertierbar -> written={written}")
        names = load_workbook(out).sheetnames
        check(names == ["Leer"], f"Platzhalter-Blatt statt Crash -> {names}")


def test_run_in_background_smoke():
    print("test_run_in_background_smoke")
    import threading as _t
    done = _t.Event()
    result = {}

    def work():
        result["ran"] = True
        done.set()

    started = et._run_in_background(work)
    check(started, "Task gestartet")
    check(done.wait(timeout=5), "work() im Hintergrund-Thread ausgeführt")
    check(result.get("ran") is True, "Seiteneffekt sichtbar")
    # Flag muss zurückgesetzt werden (kurz pollen, da finally im Thread läuft)
    for _ in range(50):
        if not et._task_running:
            break
        time.sleep(0.02)
    check(not et._task_running, "_task_running nach Abschluss zurückgesetzt")


def main():
    tests = [
        test_sanitize_sheet_name,
        test_parse_edifact,
        test_parse_orders05_xml,
        test_parse_opentrans_xml,
        test_parse_parameter_xml,
        test_split_csv_file,
        test_strip_protection_xml,
        test_entferne_schutz_integration,
        test_convert_files_to_workbook,
        test_convert_empty_creates_placeholder,
        test_run_in_background_smoke,
    ]
    for t in tests:
        t()
    print("\n" + ("=" * 50))
    if _failures:
        print(f"FEHLGESCHLAGEN: {len(_failures)} Prüfung(en)")
        for m in _failures:
            print(f"  - {m}")
        sys.exit(1)
    print("ALLE TESTS BESTANDEN")


if __name__ == "__main__":
    main()
